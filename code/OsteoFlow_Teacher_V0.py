# =============================================================
# OsteoFlow_Teacher — SVF Diffeomorphic Registration
# =============================================================

import os, re, math, sys
from pathlib import Path
from copy import deepcopy
from contextlib import contextmanager
import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import DataLoader, Dataset
import nibabel as nib
import matplotlib.pyplot as plt
import pandas as pd

try:
    from tqdm.auto import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

MODE = 'train'      # 'train' | 'inference'

BASE_DIR = Path(__file__).resolve().parent.parent
AUG_ROOT = BASE_DIR / "output_rois_augmented"
POD5_DIR = AUG_ROOT / "POD5"
POY1_DIR = AUG_ROOT / "POY1"
OUT_ROOT = BASE_DIR / "OsteoFlow_Teacher"

OUT_ROOT.mkdir(parents=True, exist_ok=True)
CKPT_DIR = OUT_ROOT / "checkpoints"
RECON_DIR = OUT_ROOT / "recons"
METRICS_DIR = OUT_ROOT / "metrics"
for d in [CKPT_DIR, RECON_DIR, METRICS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

print(f"📁 POD5: {POD5_DIR}")
print(f"📁 POY1: {POY1_DIR}")
print(f"📊 OUT:  {OUT_ROOT}")
# ----------------------- Device & Seeds -----------------------
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# ----------------------- Configuration -----------------------
ROI_SHAPE = (48, 48, 48)
HU_RANGE = (-100, 1100.0)
RANDOM_SEED = 123
torch.manual_seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)

# Train/eval
BATCH_SIZE = 1
NUM_EPOCHS = 50
LEARNING_RATE = 1e-4

# Deformation model: 'svf' (diffeomorphic via scaling-and-squaring)
DEFORMATION_MODE = 'svf'  # SVF only (diffeomorphic)

# Train/Test split (align with deform code)
SPLIT_BY_PATIENT = True  # if True, split by patient; else split by (patient, ROI)
TRAIN_SPLIT = 0.85

# Semi-online augmentation: per ROI, use original + N augmentations each epoch
USE_SEMI_ONLINE_AUG = False
# If semi-online=True: randomly pick N augmentations from aug1-aug40 each epoch
# If semi-online=False: use first N augmentations (aug1 to augN) deterministically
NUM_AUG_PER_ROI = 40  # Number of augmentations per ROI (in addition to aug0)

# UNet
UNET_BASE_CHANNELS = 48

# Classifier-Free Guidance (CFG) style conditioning
USE_CFG_DROPOUT = False  # Enable random dropout of POY1 (fixed) image during training
CFG_DROPOUT_PROB = 0.5  # Probability of dropping POY1 input (0.0 = never drop, 1.0 = always drop) - 50% forces better unconditional learning
CFG_DROPOUT_REPLACE = 'copy_x0'  # What to use when dropping POY1: 'zeros', 'noise', or 'copy_x0' (POD5)

# Loss weighting (bone emphasis)
BONE_HU_THRESHOLD = 300.0
METRICS_BONE_HU_THRESHOLD = BONE_HU_THRESHOLD  # FM V18 parity: threshold used for Dice / bone metrics
MIDDLE_SLAB_IMAGE_SLICE_START = 18  # FM V18 parity: inclusive W-slice start
MIDDLE_SLAB_IMAGE_SLICE_END = 30    # FM V18 parity: inclusive W-slice end
# Bone-weighting strength for reconstruction losses.
# When used (image-space SVF recon), per-voxel weight is:
#   w = 1 + BONE_WEIGHT_ALPHA * 1[HU_target > BONE_HU_THRESHOLD]
# so bone voxels get weight (1 + alpha) (e.g., 16x when alpha=15) and non-bone voxels get weight 1.
# This does NOT affect FM velocity MSE unless you explicitly add a bone-weighted image-space loss.
BONE_WEIGHT_ALPHA = 10.0

# SVF Diffeomorphic Registration Parameters
# φ = exp(v_svf) via scaling-and-squaring with SS_SQUARINGS iterations
SVF_FLOW_CAP_VOX = 7.0   # cap SVF velocity per-voxel per-axis (in voxels)
SS_SQUARINGS = 7         # scaling-and-squaring exponent (2^7 = 128 compositions)

# Resection plane constraint: smooth exponential weighting on deformation
# ONLY applies in IMAGE SPACE + SVF MODE (disabled for latent space or FM mode)
# At center (resection plane): weight = 0 (no deformation)
# Away from center: weight -> 1 exponentially (full deformation)
RESECTION_PLANE_CONSTRAINT = False   # If True, apply smooth exponential weighting to velocity (IMAGE + SVF only!)
RESECTION_PLANE_SIGMA = 77.0         # Controls falloff rate: smaller = sharper transition, larger = smoother
RESECTION_PLANE_PROFILE = 'gaussian'  # 'gaussian' (default), 'sigmoid', 'cosine'

# If >0, the mask will NOT go to exactly 0 at the plane.
# This often removes the visible "seam/jump" while still strongly discouraging motion.
# 0.0 keeps the original behavior (exactly zero at the center).
RESECTION_PLANE_MIN_WEIGHT = 0.0001

# Resection plane constraint for intensity residual (separate from deformation)
# ONLY applies in IMAGE SPACE + SVF MODE (FM mode has no intensity residual)
# Same smooth exponential weighting applied to additive intensity channel
INTENSITY_PLANE_CONSTRAINT = False   # If True, apply smooth exponential weighting to intensity residual (IMAGE + SVF only!)
INTENSITY_PLANE_SIGMA = 77.0         # Controls falloff rate for intensity (can differ from deformation sigma)

# Print configuration summary
print(f"\n🎯 OsteoFlow_Teacher Configuration:")
print(f"   Space: IMAGE (48³) | Deformation: SVF (diffeomorphic)")
print(f"   SVF integration: scaling-and-squaring ({SS_SQUARINGS} iterations)")
print(f"   Velocity cap: {SVF_FLOW_CAP_VOX} voxels/axis")
print(f"   UNet base channels: {UNET_BASE_CHANNELS}")
print()

# Visualization
NUM_VIS_SAMPLES = 4
SAVE_TRAIN_VIS = True  # If True, save train set visualizations each epoch (separate from test)

# ----------------------- Utils -------------------------------
def _clip_and_norm_to_unit(v_np: np.ndarray, hu_range=HU_RANGE):
    lo, hi = hu_range
    v = np.clip(v_np.astype(np.float32), lo, hi)
    v = 2.0 * (v - lo) / (hi - lo) - 1.0
    return v

def denorm_to_hu(v_np: np.ndarray, hu_range=HU_RANGE):
    lo, hi = hu_range
    return (((v_np.astype(np.float32) + 1.0) * 0.5) * (hi - lo) + lo)




def create_resection_plane_mask(shape: tuple, sigma: float = None, device: torch.device = None, profile: str = None) -> torch.Tensor:
    """Create the resection-plane weight mask (matches One_Shot_Regis.create_resection_plane_mask)."""
    if sigma is None:
        sigma = globals().get('RESECTION_PLANE_SIGMA', 3.0)

    if profile is None:
        profile = globals().get('RESECTION_PLANE_PROFILE', 'gaussian')

    if len(shape) == 5:
        B, C, D, H, W = shape
    elif len(shape) == 3:
        D, H, W = shape
        B, C = 1, 1
    else:
        raise ValueError(f"Unexpected shape: {shape}")

    sigma = max(float(sigma), 0.1)
    center_n = (W - 1) / 2.0

    w_indices = torch.arange(W, dtype=torch.float32, device=device)
    d = torch.abs(w_indices - center_n)
    x = d / sigma

    profile = str(profile).lower().strip()
    if profile == 'gaussian':
        weight_1d = 1.0 - torch.exp(-(x ** 2))
    elif profile == 'sigmoid':
        weight_1d = 2.0 * (torch.sigmoid(x) - 0.5)
    elif profile == 'cosine':
        x_clipped = torch.clamp(x, 0.0, 1.0)
        weight_1d = 0.5 * (1.0 - torch.cos(math.pi * x_clipped))
    else:
        raise ValueError(f"Unknown RESECTION_PLANE_PROFILE: {profile}. Use 'gaussian', 'sigmoid', or 'cosine'.")

    min_w = float(globals().get('RESECTION_PLANE_MIN_WEIGHT', 0.0))
    if min_w > 0.0:
        min_w = max(0.0, min(min_w, 1.0))
        weight_1d = weight_1d * (1.0 - min_w) + min_w

    mask = weight_1d.view(1, 1, W).expand(D, H, W).contiguous()
    if len(shape) == 5:
        mask = mask.unsqueeze(0).unsqueeze(0).expand(B, C, D, H, W)
    return mask

def get_resection_plane_weight_info(n_size: int, sigma: float = None, profile: str = None) -> dict:
    """Summarize the *actual* resection mask weights for the current profile.

    Notes:
    - The resection center is at (W-1)/2, which is 23.5 for W=48 (between slices).
    - For even W, the two middle slices have distance 0.5 from the plane.
    - When sigma is larger than the ROI half-width, the mask will NOT reach 1 anywhere.
    """
    if sigma is None:
        sigma = float(globals().get('RESECTION_PLANE_SIGMA', 3.0))
    if profile is None:
        profile = str(globals().get('RESECTION_PLANE_PROFILE', 'gaussian'))

    W = int(n_size)
    center_n = (W - 1) / 2.0

    # Build 1D weights on CPU for reporting (uses the same implementation as training/inference).
    mask_1d = create_resection_plane_mask((1, 1, W), sigma=sigma, device=torch.device('cpu'), profile=profile)[0, 0]

    def w_at_index(idx: int) -> float:
        idx = max(0, min(W - 1, int(idx)))
        return float(mask_1d[idx].item())

    w_min = float(mask_1d.min().item())
    w_max = float(mask_1d.max().item())
    w_edge = max(w_at_index(0), w_at_index(W - 1))

    left_mid = int(math.floor(center_n))
    right_mid = int(math.ceil(center_n))

    # For W even, these two should match (distance 0.5 to center).
    w_d0p5_left = w_at_index(left_mid)
    w_d0p5_right = w_at_index(right_mid)

    # Define approximate d=1.5 and d=2.5 by stepping outward from the middle indices.
    w_d1p5 = 0.5 * (w_at_index(left_mid - 1) + w_at_index(right_mid + 1))
    w_d2p5 = 0.5 * (w_at_index(left_mid - 2) + w_at_index(right_mid + 2))

    return {
        'center': center_n,
        'sigma': float(sigma),
        'profile': str(profile),
        'min_weight': float(globals().get('RESECTION_PLANE_MIN_WEIGHT', 0.0)),
        'w_min': w_min,
        'w_max': w_max,
        'w_edge': w_edge,
        'w_d0p5_left': w_d0p5_left,
        'w_d0p5_right': w_d0p5_right,
        'w_d1p5': float(w_d1p5),
        'w_d2p5': float(w_d2p5),
    }

def apply_resection_plane_mask(v_vox: torch.Tensor, sigma: float = None) -> torch.Tensor:
    """Apply smooth exponential resection-plane weighting (matches One_Shot_Regis.apply_resection_plane_mask)."""
    if sigma is None:
        sigma = globals().get('RESECTION_PLANE_SIGMA', 3.0)
    B, C, D, H, W = v_vox.shape
    mask = create_resection_plane_mask((B, C, D, H, W), sigma=sigma, device=v_vox.device)
    return v_vox * mask.detach()


def apply_resection_constraint_to_intensity(a_raw: torch.Tensor, sigma: float = None) -> torch.Tensor:
    """Apply the same plane weighting to intensity residual (matches One_Shot_Regis.apply_resection_constraint_to_intensity)."""
    if not globals().get('INTENSITY_PLANE_CONSTRAINT', False):
        return a_raw
    if sigma is None:
        sigma = globals().get('INTENSITY_PLANE_SIGMA', 15.0)
    B, C, D, H, W = a_raw.shape
    mask = create_resection_plane_mask((B, C, D, H, W), sigma=sigma, device=a_raw.device)
    return a_raw * mask.detach()

def _maybe_resample_to_roi(v_np, roi_shape=ROI_SHAPE):
    if tuple(v_np.shape) != tuple(roi_shape):
        raise ValueError(
        f"Input volume shape {v_np.shape} does not match expected ROI_SHAPE {roi_shape}. "
        "Please resample upstream to avoid implicit interpolation."
        )
    return v_np

def save_orthogonal_png(vol_np, save_path, title, vmin=None, vmax=None):
    D, H, W = vol_np.shape
    md, mh, mw = D//2, H//2, W//2
    if vmin is None: vmin = HU_RANGE[0]
    if vmax is None: vmax = HU_RANGE[1]
    fig, axes = plt.subplots(1, 3, figsize=(12, 4), constrained_layout=True)
    fig.suptitle(title, fontsize=14, fontweight='bold')
    axes[0].imshow(vol_np[:, mh, :].T, cmap='gray', vmin=vmin, vmax=vmax, origin='lower'); axes[0].set_title('Sagittal')
    axes[1].imshow(vol_np[md, :, :].T, cmap='gray', vmin=vmin, vmax=vmax, origin='lower'); axes[1].set_title('Coronal')
    im = axes[2].imshow(vol_np[:, :, mw], cmap='gray', vmin=vmin, vmax=vmax, origin='lower'); axes[2].set_title('Axial')
    fig.colorbar(im, ax=axes.ravel().tolist(), shrink=0.8, label='HU')
    for ax in axes: ax.set_xticks([]); ax.set_yticks([])
    plt.savefig(save_path, dpi=150); plt.close()

def save_combined_comparison(input_hu, gt_hu, pred_hu, save_path, epoch, case_info="", pred_label="POY1 Prediction",
                    cfg_uncond_hu=None, cfg_cond_hu=None, show_resection_mask=None):
    from matplotlib.colors import Normalize
    error_hu = np.abs(pred_hu - gt_hu)
    
    # Check if resection mask should be shown
    if show_resection_mask is None:
        show_resection_mask = globals().get('RESECTION_PLANE_CONSTRAINT', False)
    
    # Base volumes (always present)
    volumes = [
        (input_hu, "POD5 Input (Pre-op)", 'gray', HU_RANGE[0], HU_RANGE[1], 'HU'),
        (gt_hu, "POY1 Ground Truth (Post-op)", 'gray', HU_RANGE[0], HU_RANGE[1], 'HU'),
        (pred_hu, f"{pred_label} (Epoch {epoch})", 'gray', HU_RANGE[0], HU_RANGE[1], 'HU'),
        (error_hu, "|Pred - GT| Error", 'hot', 0, 200, 'HU Error')
    ]
    
    # Add CFG rows if provided
    if cfg_uncond_hu is not None and cfg_cond_hu is not None:
        volumes.extend([
        (cfg_uncond_hu, "CFG Unconditional (POD5-only)", 'gray', HU_RANGE[0], HU_RANGE[1], 'HU'),
        (cfg_cond_hu, "CFG Conditional (with POY1)", 'gray', HU_RANGE[0], HU_RANGE[1], 'HU')
        ])
    
    # Add resection mask row if enabled
    if show_resection_mask:
        D, H, W = pred_hu.shape
        sigma = globals().get('RESECTION_PLANE_SIGMA', 3.0)
        mask_np = create_resection_plane_mask((D, H, W), sigma=sigma).numpy()
        # Invert for visualization: high value = protected (low deform), low = free to deform
        mask_vis = 1.0 - mask_np  # Now: 1 = center (no deform), 0 = edges (full deform)
        info = get_resection_plane_weight_info(W, sigma, profile=globals().get('RESECTION_PLANE_PROFILE', 'gaussian'))
        mask_label = (
        f"Resection Weight (profile={info['profile']}, σ={sigma:.1f}, center={info['center']:.1f}, "
        f"w@0.5={info['w_d0p5_left']:.3f}, edge={info['w_edge']:.3f})"
        )
        volumes.append((mask_vis, mask_label, 'Reds', 0, 1, 'Protection'))
    
    num_rows = len(volumes)
    fig_height = 4 * num_rows
    fig, axes = plt.subplots(num_rows, 3, figsize=(15, fig_height), constrained_layout=True)
    title = f"Training Progress - Epoch {epoch}"
    if case_info: title += f" | {case_info}"
    fig.suptitle(title, fontsize=20, fontweight='bold')
    for row_idx, (vol_np, row_title, cmap, vmin, vmax, cbar_label) in enumerate(volumes):
        D, H, W = vol_np.shape
        md, mh, mw = D//2, H//2, W//2
        norm = Normalize(vmin=vmin, vmax=vmax)
        axes[row_idx, 0].imshow(vol_np[:, mh, :].T, cmap=cmap, norm=norm, origin='lower')
        axes[row_idx, 0].set_title('Sagittal' if row_idx == 0 else '', fontsize=14, fontweight='bold')
        axes[row_idx, 0].set_ylabel(row_title, fontsize=13, fontweight='bold')
        axes[row_idx, 0].set_xticks([]); axes[row_idx, 0].set_yticks([])
        axes[row_idx, 1].imshow(vol_np[md, :, :].T, cmap=cmap, norm=norm, origin='lower')
        axes[row_idx, 1].set_title('Coronal' if row_idx == 0 else '', fontsize=14, fontweight='bold')
        axes[row_idx, 1].set_xticks([]); axes[row_idx, 1].set_yticks([])
        im = axes[row_idx, 2].imshow(vol_np[:, :, mw], cmap=cmap, norm=norm, origin='lower')
        axes[row_idx, 2].set_title('Axial' if row_idx == 0 else '', fontsize=14, fontweight='bold')
        axes[row_idx, 2].set_xticks([]); axes[row_idx, 2].set_yticks([])
        cbar = plt.colorbar(im, ax=axes[row_idx, :], shrink=0.8, pad=0.02)
        cbar.set_label(cbar_label, rotation=270, labelpad=20, fontsize=12, fontweight='bold')
        cbar.ax.tick_params(labelsize=10)
    plt.savefig(save_path, dpi=150, bbox_inches='tight'); plt.close()

# ----------------------- Differentiable Warper & SVF exp -----------------------
def _make_base_grid(D, H, W, device):
    # grid_sample expects last dim order = (x, y, z) in normalized [-1,1]
    z = torch.linspace(-1, 1, D, device=device)
    y = torch.linspace(-1, 1, H, device=device)
    x = torch.linspace(-1, 1, W, device=device)
    zz, yy, xx = torch.meshgrid(z, y, x, indexing='ij')
    grid = torch.stack([xx, yy, zz], dim=-1)  # [D,H,W,3] as (x,y,z)
    return grid

def _vox2norm_displacement(vox_disp: torch.Tensor, apply_resection_mask: bool = None) -> torch.Tensor:
    """
    Convert voxel-unit displacement (channels z,y,x) [B,3,D,H,W]
    into normalized displacement field [B,D,H,W,3] ordered as (x,y,z).
    
    Args:
        vox_disp: velocity/displacement field [B, 3, D, H, W]
        apply_resection_mask: if True, zero velocity in resection plane region (uses global flag if None)
    """
    # Apply resection plane mask if enabled
    if apply_resection_mask is None:
        apply_resection_mask = globals().get('RESECTION_PLANE_CONSTRAINT', False)
    if apply_resection_mask:
        sigma_val = globals().get('RESECTION_PLANE_SIGMA', 3.0)
        vox_disp = apply_resection_plane_mask(vox_disp, sigma=sigma_val)
    
    B, C, D, H, W = vox_disp.shape
    assert C == 3
    dz = vox_disp[:, 0]  # [B,D,H,W]
    dy = vox_disp[:, 1]
    dx = vox_disp[:, 2]
    # Convert vox -> normalized [-1,1] deltas (per axis uses size-1)
    nx = 2.0 * dx / max(W - 1, 1)
    ny = 2.0 * dy / max(H - 1, 1)
    nz = 2.0 * dz / max(D - 1, 1)
    disp = torch.stack([nx, ny, nz], dim=-1)  # [B,D,H,W,3] (x,y,z)
    return disp

def _sample_field(field_xyz: torch.Tensor, grid_xyzn: torch.Tensor) -> torch.Tensor:
    """
    field_xyz: [B,D,H,W,3] normalized displacement or vector field (x,y,z order in last dim)
    grid_xyzn: [B,D,H,W,3] normalized sampling grid
    returns sampled field at grid: [B,D,H,W,3]
    """
    B, D, H, W, _ = field_xyz.shape
    # grid_sample expects channels-first; permute:
    field_ch_first = field_xyz.permute(0, 4, 1, 2, 3)  # [B,3,D,H,W]
    sampled = F.grid_sample(field_ch_first, grid_xyzn, mode='bilinear',
                    padding_mode='border', align_corners=True)
    sampled = sampled.permute(0, 2, 3, 4, 1)  # back to [B,D,H,W,3]
    return sampled

def compose_fields(phi_a: torch.Tensor, phi_b: torch.Tensor) -> torch.Tensor:
    """
    Compose two normalized displacement fields (x,y,z last dim), both on same grid:
      result = phi_b + phi_a o (Id + phi_b)
    where composition uses sampling of phi_a at coordinates (Id + phi_b).
    Shapes: [B,D,H,W,3]
    """
    B, D, H, W, _ = phi_a.shape
    device = phi_a.device
    base = _make_base_grid(D, H, W, device).unsqueeze(0).repeat(B, 1, 1, 1, 1)  # [B,D,H,W,3]
    grid_b = base + phi_b
    phi_a_warped = _sample_field(phi_a, grid_b)
    return phi_b + phi_a_warped

def expv_scaling_squaring(v_vox: torch.Tensor, n_squarings: int = SS_SQUARINGS, apply_resection_mask: bool = None) -> torch.Tensor:
    """
    Exponentiate stationary velocity field v (voxel units, channels=z,y,x, shape [B,3,D,H,W])
    into a diffeomorphic displacement field phi (normalized coords, [B,D,H,W,3])
    using scaling-and-squaring in normalized coordinate space.
    
    Args:
        v_vox: velocity field [B, 3, D, H, W]
        n_squarings: number of squaring iterations
        apply_resection_mask: if True, zero velocity in resection plane region (uses global flag if None)
    """
    # Apply resection plane mask if enabled
    if apply_resection_mask is None:
        apply_resection_mask = globals().get('RESECTION_PLANE_CONSTRAINT', False)
    if apply_resection_mask:
        sigma_val = globals().get('RESECTION_PLANE_SIGMA', 3.0)
        v_vox = apply_resection_plane_mask(v_vox, sigma=sigma_val)
    
    # Read SVF cap from globals at runtime (allows _temp_global_overrides to work for preprocessing)
    svf_cap = float(globals().get('SVF_FLOW_CAP_VOX', SVF_FLOW_CAP_VOX))
    # Cap v for stability (per-axis, per-voxel)
    v_vox_capped = svf_cap * torch.tanh(v_vox / max(svf_cap, 1e-6))
    # Convert to normalized displacement per unit time (don't apply mask again - already applied above)
    v_norm = _vox2norm_displacement(v_vox_capped, apply_resection_mask=False)  # [B,D,H,W,3]
    # Scale down
    phi = v_norm / (2.0 ** n_squarings)
    # Repeated squaring (phi = phi ∘ phi)
    for _ in range(n_squarings):
        phi = compose_fields(phi, phi)
    return phi  # normalized displacement field



def warp_image_with_phi_norm(x: torch.Tensor, phi_norm: torch.Tensor) -> torch.Tensor:
    """
    Warp image x by normalized displacement field phi (both on same grid).
    x: [B,1,D,H,W] in [-1,1]
    phi_norm: [B,D,H,W,3] normalized displacements (x,y,z order)
    """
    B, C, D, H, W = x.shape
    base = _make_base_grid(D, H, W, x.device).unsqueeze(0).repeat(B, 1, 1, 1, 1)
    grid = base + phi_norm  # [B,D,H,W,3]
    x_warp = F.grid_sample(x, grid, mode='bilinear', padding_mode='border', align_corners=True)
    return x_warp

def _rigid_grid(D: int, H: int, W: int, device: torch.device, angles_t: torch.Tensor, trans_t: torch.Tensor) -> torch.Tensor:
    """Build normalized sampling grid for rigid transform via affine_grid (fully differentiable)."""
    B = angles_t.shape[0]
    ax, ay, az = angles_t[:, 0], angles_t[:, 1], angles_t[:, 2]
    cx, sx = torch.cos(ax), torch.sin(ax)
    cy, sy = torch.cos(ay), torch.sin(ay)
    cz, sz = torch.cos(az), torch.sin(az)
    R11 = cz * cy
    R12 = cz * sy * sx - sz * cx
    R13 = cz * sy * cx + sz * sx
    R21 = sz * cy
    R22 = sz * sy * sx + cz * cx
    R23 = sz * sy * cx - cz * sx
    R31 = -sy
    R32 = cy * sx
    R33 = cy * cx
    R = torch.stack([
        torch.stack([R11, R12, R13], dim=-1),
        torch.stack([R21, R22, R23], dim=-1),
        torch.stack([R31, R32, R33], dim=-1)
    ], dim=-2)  # [B,3,3]
    # Build theta without in-place ops to keep gradients
    theta = torch.cat([R, trans_t.unsqueeze(-1)], dim=-1)  # [B,3,4]
    size = (B, 1, D, H, W)
    grid = F.affine_grid(theta, size=size, align_corners=True)
    return grid


def rigid_register(x0: torch.Tensor, x1: torch.Tensor, iters: int = 100, lr: float = 1e-2) -> tuple[torch.Tensor, torch.Tensor, torch.Tensor]:
    """
    Estimate rigid transform aligning x0->x1 by optimizing angles+translation.
    This is inference-only (no UNet training).
    Returns: x0_warped, angles, translations.
    """
    B, C, D, H, W = x0.shape
    device = x0.device
    # Enable gradients locally even if called from torch.no_grad() context
    with torch.enable_grad():
        angles = torch.zeros(B, 3, device=device, requires_grad=True)
        trans = torch.zeros(B, 3, device=device, requires_grad=True)
        opt = torch.optim.Adam([angles, trans], lr=lr)
        x0_detach = x0.detach()  # Detach inputs so we only optimize angles/trans
        x1_detach = x1.detach()
        for _ in range(iters):
            opt.zero_grad()
            grid = _rigid_grid(D, H, W, device, angles, trans)
            x0_warp = F.grid_sample(x0_detach, grid, mode='bilinear', padding_mode='border', align_corners=True)
            loss = F.mse_loss(x0_warp, x1_detach)
            loss.backward()
            opt.step()
    with torch.no_grad():
        grid = _rigid_grid(D, H, W, device, angles, trans)
        x0_warp = F.grid_sample(x0, grid, mode='bilinear', padding_mode='border', align_corners=True)
    return x0_warp, angles.detach(), trans.detach()


def affine_register(x0: torch.Tensor, x1: torch.Tensor, iters: int = 150, lr: float = 1e-2) -> tuple[torch.Tensor, torch.Tensor]:
    """
    Estimate affine transform (12-DOF) aligning x0->x1 by directly optimizing the affine matrix.
    Parameterization: 3x4 matrix [R|t] where R includes rotation, scale, and shear.
    This is inference-only (no UNet training).
    Returns: x0_warped, affine_matrix [B, 3, 4].
    """
    B, C, D, H, W = x0.shape
    device = x0.device
    # Enable gradients locally even if called from torch.no_grad() context
    with torch.enable_grad():
        # Initialize to identity: [I | 0]
        affine_matrix = torch.eye(3, 4, device=device).unsqueeze(0).repeat(B, 1, 1).requires_grad_(True)
        opt = torch.optim.Adam([affine_matrix], lr=lr)
        x0_detach = x0.detach()
        x1_detach = x1.detach()
        for _ in range(iters):
            opt.zero_grad()
            grid = F.affine_grid(affine_matrix, size=(B, C, D, H, W), align_corners=True)
            x0_warp = F.grid_sample(x0_detach, grid, mode='bilinear', padding_mode='border', align_corners=True)
            loss = F.mse_loss(x0_warp, x1_detach)
            loss.backward()
            opt.step()
    with torch.no_grad():
        grid = F.affine_grid(affine_matrix, size=(B, C, D, H, W), align_corners=True)
        x0_warp = F.grid_sample(x0, grid, mode='bilinear', padding_mode='border', align_corners=True)
    return x0_warp, affine_matrix.detach()


def jacobian_determinant(phi_norm: torch.Tensor) -> torch.Tensor:
    """
    Compute Jacobian determinant of the deformation map T(x) = x + phi(x) in normalized coords.
    phi_norm: [B, D, H, W, 3] with components ordered (x, y, z) in last dim.
    Returns: detJ tensor [B, D, H, W].
    """
    # Ensure last dim order is (x, y, z)
    # Finite differences with central scheme; replicate padding at borders
    B, D, H, W, C = phi_norm.shape
    assert C == 3, "phi_norm must have 3 components (x,y,z)"
    # Compute spatial gradients of each component
    def grad_component(u):
        # u: [B, D, H, W]
        ux = (F.pad(u, (1,1,0,0,0,0), mode='replicate')[:, :, :, 2:] - F.pad(u, (1,1,0,0,0,0), mode='replicate')[:, :, :, :-2]) * 0.5
        uy = (F.pad(u, (0,0,1,1,0,0), mode='replicate')[:, :, 2:, :] - F.pad(u, (0,0,1,1,0,0), mode='replicate')[:, :, :-2, :]) * 0.5
        uz = (F.pad(u, (0,0,0,0,1,1), mode='replicate')[:, 2:, :, :] - F.pad(u, (0,0,0,0,1,1), mode='replicate')[:, :-2, :, :]) * 0.5
        return ux, uy, uz
    # Split components
    phi_x = phi_norm[..., 0].contiguous()
    phi_y = phi_norm[..., 1].contiguous()
    phi_z = phi_norm[..., 2].contiguous()
    # Gradients
    dxx, dxy, dxz = grad_component(phi_x)
    dyx, dyy, dyz = grad_component(phi_y)
    dzx, dzy, dzz = grad_component(phi_z)
    # Jacobian = I + grad(phi)
    J11 = 1.0 + dxx
    J12 = dxy
    J13 = dxz
    J21 = dyx
    J22 = 1.0 + dyy
    J23 = dyz
    J31 = dzx
    J32 = dzy
    J33 = 1.0 + dzz
    # Determinant of 3x3
    detJ = (
        J11 * (J22 * J33 - J23 * J32)
        - J12 * (J21 * J33 - J23 * J31)
        + J13 * (J21 * J32 - J22 * J31)
    )
    return detJ

# NOTE: integrate_rectified_flow function removed - this module now uses SVF-only mode
# For diffeomorphic integration, use expv_scaling_squaring() instead

# ----------------------- Time Embedding (FM V10 Style) -----------------------
class TimeEmbedding(nn.Module):
    """Sinusoidal time embedding with learned projection (FM V10 style)."""
    def __init__(self, d_t=64, M=1000, out_channels=128):
        super().__init__()
        self.d_t = d_t
        self.M = M
        self.proj = nn.Sequential(
        nn.Linear(d_t, out_channels * 4),
        nn.GELU(),
        nn.Linear(out_channels * 4, out_channels)
        )
    
    def forward(self, t):
        # t: [B] scalar time values in [0,1]
        import math
        h = self.d_t // 2
        freqs = torch.exp(-math.log(self.M) * torch.arange(h, device=t.device) / max(h-1, 1))
        ang = (t[:, None] * self.M) * freqs[None]
        emb = torch.cat([ang.sin(), ang.cos()], dim=-1)
        if self.d_t % 2:
            emb = F.pad(emb, (0, 1))
        return self.proj(emb)  # [B, out_channels]

# ----------------------- Dataset -----------------------------
class ROI3DDataset(Dataset):
    """
    Pair POD5 and POY1 ROIs. Enforce 48x48x48. Provide both normalized [-1,1] and raw HU (for weighting).
    Matches your robust filename pairing.
    """
    def __init__(self, pod5_dir, poy1_dir, normalize=True, recursive=False):
        self.pod5_dir = Path(pod5_dir)
        self.poy1_dir = Path(poy1_dir)
        self.normalize = normalize
        self.pairs = []

        pod5_pat = re.compile(
            r'^(?P<case>\d+)_POD5_(?P<roi>ROI\d+)(?:_(?P<aug>aug\d+))?(?:_RAS)?\.nii(?:\.gz)?$',
            re.IGNORECASE
        )
        globber = self.pod5_dir.rglob if recursive else self.pod5_dir.glob
        pod5_files = list(globber("*.nii")) + list(globber("*.nii.gz"))

        def first_existing(cands):
            for c in cands:
                if c.exists(): return c
            return None

        for pod5_file in pod5_files:
            m = pod5_pat.match(pod5_file.name)
            if not m: continue
            case_str = m.group("case"); roi_str = m.group("roi"); aug = m.group("aug")
            aug_id = int(aug.replace('aug','')) if aug else 0
            name_roots = []
            if aug:
                name_roots += [f"{case_str}_POY1_{roi_str}_{aug}_RAS", f"{case_str}_POY1_{roi_str}_{aug}"]
            name_roots += [f"{case_str}_POY1_{roi_str}_RAS", f"{case_str}_POY1_{roi_str}"]
            candidates = []
            for root in name_roots:
                candidates += [self.poy1_dir / f"{root}.nii.gz", self.poy1_dir / f"{root}.nii"]
            target = first_existing(candidates)
            if target is None: continue
            self.pairs.append({
                "pod5_path": pod5_file,
                "poy1_path": target,
                "case_id": int(case_str),
                "roi_num": int(roi_str.replace('ROI','')),
                "aug_id": aug_id,
            })
        self.pairs.sort(key=lambda p: (p["case_id"], p["roi_num"], str(p["pod5_path"])) )
        print(f"📊 Found {len(self.pairs)} paired ROI volumes")

    def __len__(self): return len(self.pairs)

    def _load_np(self, path: Path):
        img = nib.load(str(path))
        v = img.get_fdata().astype(np.float32)
        v = _maybe_resample_to_roi(v, ROI_SHAPE)
        return v

    def __getitem__(self, idx):
        p = self.pairs[idx]
        x0_hu = self._load_np(p["pod5_path"])  # POD5 raw HU
        x1_hu = self._load_np(p["poy1_path"])  # POY1 raw HU
        x0 = _clip_and_norm_to_unit(x0_hu, HU_RANGE)
        x1 = _clip_and_norm_to_unit(x1_hu, HU_RANGE)
        
        x0_tensor = torch.from_numpy(x0).unsqueeze(0)  # [1,D,H,W]
        
        return {
        "x0": x0_tensor,   # [1,D,H,W]
        "x1": torch.from_numpy(x1).unsqueeze(0),
        "x0_hu": torch.from_numpy(x0_hu).unsqueeze(0),
        "x1_hu": torch.from_numpy(x1_hu).unsqueeze(0),
        "meta": {
        "case_id": p["case_id"],
        "roi_num": p["roi_num"],
        "aug_id": p.get("aug_id", 0),
        "pod5_path": str(p["pod5_path"]),
        "poy1_path": str(p["poy1_path"]),
        }
        }


class ResBlock3DPlain(nn.Module):
    def __init__(self, ch, pdrop=0.3):
        super().__init__()
        self.n1 = nn.GroupNorm(min(8, ch), ch)
        self.c1 = nn.Conv3d(ch, ch, 3, padding=1)
        self.n2 = nn.GroupNorm(min(8, ch), ch)
        self.d = nn.Dropout3d(pdrop)
        self.c2 = nn.Conv3d(ch, ch, 3, padding=1)
    def forward(self, x):
        h = self.c1(F.silu(self.n1(x)))
        h = self.c2(self.d(F.silu(self.n2(h))))
        return x + h

class ResBlock3DWithTimeEmb(nn.Module):
    """Residual block with time embedding modulation (FM V10 style)."""
    def __init__(self, channels, emb_channels, dropout=0.1):
        super().__init__()
        self.norm1 = nn.GroupNorm(min(8, channels), channels)
        self.conv1 = nn.Conv3d(channels, channels, 3, padding=1)
        
        # Time embedding projection (scale and shift)
        self.emb_proj = nn.Sequential(
        nn.SiLU(),
        nn.Linear(emb_channels, channels * 2)
        )
        
        self.norm2 = nn.GroupNorm(min(8, channels), channels)
        self.dropout = nn.Dropout3d(dropout)
        self.conv2 = nn.Conv3d(channels, channels, 3, padding=1)
    
    def forward(self, x, emb):
        """
        x: [B, C, D, H, W]
        emb: [B, emb_channels] time embedding
        """
        h = self.conv1(F.silu(self.norm1(x)))
        
        # Apply time modulation (scale and shift)
        emb_out = self.emb_proj(emb)[:, :, None, None, None]  # [B, 2*C, 1, 1, 1]
        scale, shift = emb_out.chunk(2, dim=1)
        h = h * (1 + scale) + shift
        
        h = self.conv2(self.dropout(F.silu(self.norm2(h))))
        return x + h

class UNet3D(nn.Module):
    """
    TEACHER UNet for diffeomorphic registration via SVF.
    
    IMPORTANT: This is the TEACHER model - outputs VELOCITY ONLY (3 channels).
    The preprocessing model outputs velocity + intensity (4 channels).
    
    This network learns a Stationary Velocity Field (SVF) that induces diffeomorphic
    transformations via scaling-and-squaring: φ = exp(v_svf).
    
    Mathematical Foundation:
        - v_svf(ξ): stationary velocity field over spatial coordinates ξ∈Ω
        - φ_t = exp(t·v_svf): one-parameter family of diffeomorphisms
        - x̃_t = x^(POD5) ∘ φ_t: intermediate anatomies (pure deformation)
    
    IMAGE SPACE mode:
        Input: concat(x0, x1) [B, 2, 48, 48, 48] -> outputs [B, 3, 48, 48, 48]
        Spatial: 48 -> 24 -> 12 -> 24 -> 48 (2 levels, matching FM_V26)
    
    Output:
        - (vz, vy, vx) = SVF in voxel units, capped by SVF_FLOW_CAP_VOX
        - NO intensity residual (pure deformation)
    
    VoxelMorph-style: network sees both moving (POD5) and fixed (POY1) images.
    """
    def __init__(self, base_channels=UNET_BASE_CHANNELS, use_time_emb=False):
        super().__init__()
        c = base_channels
        
        self.use_time_emb = use_time_emb
        
        # Time embedding for FM mode (FM V10 style)
        if use_time_emb:
            emb_channels = 256
            self.time_emb = TimeEmbedding(d_t=64, M=1000, out_channels=emb_channels)
            self.emb_channels = emb_channels
            print(f"   ⏱️  Time embedding enabled (emb_channels={emb_channels})")
        else:
            self.emb_channels = None
        
        # Input channels: 2 images concatenated (+ optional cond flag)
        img_ch = 1
        in_ch = 2 * img_ch + (1 if globals().get('USE_CFG_DROPOUT', False) else 0)
        self._expect_cond_flag = (in_ch == 2 * img_ch + 1)
        self.in_conv = nn.Conv3d(in_ch, c, 3, padding=1)
        
        # IMAGE SPACE: 48³ input -> 2 downsampling levels (48->24->12) [FM_V26 architecture]
        # enc - use time-aware blocks if time embedding enabled
        if use_time_emb:
            self.e1_1 = ResBlock3DWithTimeEmb(c, emb_channels);   self.e1_2 = ResBlock3DWithTimeEmb(c, emb_channels)
            self.e2_1 = ResBlock3DWithTimeEmb(c*2, emb_channels); self.e2_2 = ResBlock3DWithTimeEmb(c*2, emb_channels)
            self.b1 = ResBlock3DWithTimeEmb(c*4, emb_channels);   self.b2 = ResBlock3DWithTimeEmb(c*4, emb_channels)
            self.d2_1 = ResBlock3DWithTimeEmb(c*4, emb_channels); self.d2_2 = ResBlock3DWithTimeEmb(c*4, emb_channels)
            self.d1_1 = ResBlock3DWithTimeEmb(c*2, emb_channels); self.d1_2 = ResBlock3DWithTimeEmb(c*2, emb_channels)
        else:
            self.e1_1 = ResBlock3DPlain(c);   self.e1_2 = ResBlock3DPlain(c)
            self.e2_1 = ResBlock3DPlain(c*2); self.e2_2 = ResBlock3DPlain(c*2)
            self.b1 = ResBlock3DPlain(c*4);   self.b2 = ResBlock3DPlain(c*4)
            self.d2_1 = ResBlock3DPlain(c*4); self.d2_2 = ResBlock3DPlain(c*4)
            self.d1_1 = ResBlock3DPlain(c*2); self.d1_2 = ResBlock3DPlain(c*2)
        self.down1 = nn.Conv3d(c, c*2, 3, stride=2, padding=1)   # 48->24
        self.down2 = nn.Conv3d(c*2, c*4, 3, stride=2, padding=1) # 24->12
        # dec
        self.up2 = nn.ConvTranspose3d(c*4, c*2, 3, stride=2, padding=1, output_padding=1) # 12->24
        self.p2 = nn.Conv3d(c*4, c*2, 1)
        self.up1 = nn.ConvTranspose3d(c*2, c,   3, stride=2, padding=1, output_padding=1) # 24->48
        self.p1 = nn.Conv3d(c*2, c, 1)
        # Dummy layers for compatibility (not used in 2-level mode)
        self.e3_1 = self.e3_2 = self.down3 = self.up3 = self.d3_1 = self.d3_2 = self.p3 = None
        # Output head: SVF mode - VELOCITY ONLY (vz,vy,vx) -> 3 channels
        out_ch = 3
        self.out = nn.Sequential(nn.GroupNorm(min(8, c), c), nn.SiLU(), nn.Conv3d(c, out_ch, 3, padding=1))
        # Initialize last layer small to start near identity
        nn.init.zeros_(self.out[-1].weight); nn.init.zeros_(self.out[-1].bias)
        # Rigid 6-DOF head (angles rx,ry,rz in radians; translations tx,ty,tz in voxels)
        self.rigid_head = nn.Sequential(
            nn.GroupNorm(min(8, c), c),
            nn.SiLU(),
            nn.Conv3d(c, 6, 1, padding=0)
        )
        # Initialize rigid head to output zeros -> identity transform at start
        nn.init.zeros_(self.rigid_head[-1].weight)
        nn.init.zeros_(self.rigid_head[-1].bias)

    def forward(self, x0, x1, cond_flag: torch.Tensor | None = None, t: torch.Tensor | None = None):
        # Concatenate moving (x0) and fixed (x1) like VoxelMorph
        # Optionally include a conditioning-available flag channel (0=uncond, 1=cond)
        # In FM mode with time embedding: t is used instead of cond_flag
        if (cond_flag is not None) and self._expect_cond_flag:
            x = torch.cat([x0, x1, cond_flag], dim=1)
        else:
            x = torch.cat([x0, x1], dim=1)  # [B, 2*C, D, H, W]
        h0 = self.in_conv(x)
        
        # Get time embedding ONLY if using FM mode with time conditioning enabled
        # For SVF mode or FM without t conditioning: t_emb stays None, plain ResBlocks used
        t_emb = None
        if self.use_time_emb and t is not None:
            t_emb = self.time_emb(t)  # [B, emb_channels]
        
        # IMAGE SPACE: 48³ -> 24³ -> 12³ -> 24³ -> 48³ (2 levels, FM_V26 architecture)
        if self.use_time_emb and t_emb is not None:
            # Time-modulated blocks (FM V10 style)
            h1 = self.e1_1(h0, t_emb); h1 = self.e1_2(h1, t_emb); s1 = h1
            h2 = self.down1(h1)
            h2 = self.e2_1(h2, t_emb); h2 = self.e2_2(h2, t_emb); s2 = h2
            hb = self.down2(h2)
            hb = self.b1(hb, t_emb); hb = self.b2(hb, t_emb)
            u2 = self.up2(hb); u2 = torch.cat([u2, s2], dim=1)
            u2 = self.d2_1(u2, t_emb); u2 = self.d2_2(u2, t_emb); u2 = self.p2(u2)
            u1 = self.up1(u2); u1 = torch.cat([u1, s1], dim=1)
            u1 = self.d1_1(u1, t_emb); u1 = self.d1_2(u1, t_emb); u1 = self.p1(u1)
        else:
            # Plain blocks (SVF or FM without t conditioning)
            h1 = self.e1_1(h0); h1 = self.e1_2(h1); s1 = h1
            h2 = self.down1(h1)
            h2 = self.e2_1(h2); h2 = self.e2_2(h2); s2 = h2
            hb = self.down2(h2)
            hb = self.b1(hb); hb = self.b2(hb)
            u2 = self.up2(hb); u2 = torch.cat([u2, s2], dim=1)
            u2 = self.d2_1(u2); u2 = self.d2_2(u2); u2 = self.p2(u2)
            u1 = self.up1(u2); u1 = torch.cat([u1, s1], dim=1)
            u1 = self.d1_1(u1); u1 = self.d1_2(u1); u1 = self.p1(u1)
        
        out = self.out(u1)  # [B, 3, D, H, W]
        
        # SVF mode: output VELOCITY ONLY (3 channels)
        # v_out: (vz, vy, vx) in voxels, capped by SVF_FLOW_CAP_VOX via tanh
        v_out = out  # (z,y,x) in voxels - all 3 channels
        v_out = SVF_FLOW_CAP_VOX * torch.tanh(v_out / max(SVF_FLOW_CAP_VOX, 1e-6))
        
        return v_out

    def forward_rigid(self, x0, x1, cond_flag: torch.Tensor | None = None):
        # Same encoder-decoder path to get features, then predict global 6-DOF
        if (cond_flag is not None) and self._expect_cond_flag:
            x = torch.cat([x0, x1, cond_flag], dim=1)
        else:
            x = torch.cat([x0, x1], dim=1)
        h0 = self.in_conv(x)
        
        # IMAGE SPACE: 48³ -> 24³ -> 12³ -> 24³ -> 48³ (2 levels, FM_V26 architecture)
        h1 = self.e1_1(h0); h1 = self.e1_2(h1); s1 = h1
        h2 = self.down1(h1)
        h2 = self.e2_1(h2); h2 = self.e2_2(h2); s2 = h2
        hb = self.down2(h2)
        hb = self.b1(hb); hb = self.b2(hb)
        u2 = self.up2(hb); u2 = torch.cat([u2, s2], dim=1)
        u2 = self.d2_1(u2); u2 = self.d2_2(u2); u2 = self.p2(u2)
        u1 = self.up1(u2); u1 = torch.cat([u1, s1], dim=1)
        u1 = self.d1_1(u1); u1 = self.d1_2(u1); u1 = self.p1(u1)
        
        rigid_field = self.rigid_head(u1)  # [B,6,D,H,W]
        # Global average pooling to parameters [B,6]
        angles_trans = rigid_field.mean(dim=(2,3,4))
        rx, ry, rz, tx, ty, tz = torch.split(angles_trans, 1, dim=1)
        # Constrain angles to small rotations (tanh * max_angle) and translations (tanh * max_trans)
        max_angle = 0.1  # ~5.7 degrees max rotation
        max_trans = 0.1  # max 10% of normalized coord range
        rx = max_angle * torch.tanh(rx)
        ry = max_angle * torch.tanh(ry)
        rz = max_angle * torch.tanh(rz)
        tx = max_trans * torch.tanh(tx)
        ty = max_trans * torch.tanh(ty)
        tz = max_trans * torch.tanh(tz)
        return rx.squeeze(1), ry.squeeze(1), rz.squeeze(1), tx.squeeze(1), ty.squeeze(1), tz.squeeze(1)


class SVFDiffeomorphicTeacher(nn.Module):
    """
    Diffeomorphic Teacher using Stationary Velocity Field (SVF).
    
    This wrapper is used by FM student models for distillation. It:
    1. Takes (POD5, POY1) image pairs in image space [B, 1, 48, 48, 48]
    2. Uses the frozen SVF UNet to predict v_svf(POD5, POY1)
    3. Computes intermediate states via diffeomorphic path: φ_t = exp(t * v_svf)
    4. Returns finite-difference velocity in IMAGE space for teacher signal
    
    Mathematical Foundation:
        v_svf(ξ): stationary velocity field over spatial coordinates ξ∈Ω
        φ_t = exp(t · v_svf): diffeomorphic warp at time t ∈ [0, 1]
        x̃_t = POD5 ∘ φ_t: intermediate anatomy at time t
        v_teacher(t) = (x̃_{t+ε} - x̃_t) / ε: finite difference velocity in image space
    """
    
    def __init__(self, svf_model: nn.Module, fd_epsilon: float = 0.01):
        super().__init__()
        self.svf_model = svf_model
        self.fd_epsilon = fd_epsilon
        
        # Freeze the SVF model
        self.svf_model.eval()
        for p in self.svf_model.parameters():
            p.requires_grad_(False)
    
    @torch.no_grad()
    def get_svf(self, pod5_img: torch.Tensor, poy1_img: torch.Tensor) -> torch.Tensor:
        """
        Get the SVF velocity field from the frozen SVF model.
        
        Args:
            pod5_img: POD5 image [B, 1, 48, 48, 48] normalized to [-1, 1]
            poy1_img: POY1 image [B, 1, 48, 48, 48] normalized to [-1, 1]
        
        Returns:
            v_svf: Velocity field [B, 3, 48, 48, 48] in voxel units
        """
        # Compute conditioning flag (1 if images differ, 0 otherwise)
        cond_flag = (poy1_img != pod5_img).float().mean(dim=1, keepdim=True)
        
        # Get SVF from model (returns v_vox, a_raw)
        v_svf = self.svf_model(pod5_img, poy1_img, cond_flag=cond_flag, t=None)
        return v_svf
    
    @torch.no_grad()
    def get_warped_at_t(self, pod5_img: torch.Tensor, v_svf: torch.Tensor, t: torch.Tensor) -> torch.Tensor:
        """
        Warp POD5 image using φ_t = exp(t * v_svf).
        
        Args:
            pod5_img: POD5 image [B, 1, 48, 48, 48]
            v_svf: SVF velocity [B, 3, 48, 48, 48]
            t: Time values [B] in [0, 1]
        
        Returns:
            x_t: Warped image at time t [B, 1, 48, 48, 48]
        """
        B = pod5_img.size(0)
        
        # Scale velocity by t: v_t = t * v_svf
        t_expanded = t.view(B, 1, 1, 1, 1)
        v_t = v_svf * t_expanded
        
        # Compute diffeomorphic warp: φ_t = exp(v_t) via scaling-and-squaring
        phi_norm = expv_scaling_squaring(v_t, n_squarings=SS_SQUARINGS, apply_resection_mask=False)
        
        # Warp POD5 image
        x_t = warp_image_with_phi_norm(pod5_img, phi_norm)
        return x_t
    
    @torch.no_grad()
    def get_teacher_velocity_at_t(
        self,
        pod5_img: torch.Tensor,
        poy1_img: torch.Tensor,
        t: torch.Tensor
    ) -> torch.Tensor:
        """
        Compute teacher velocity in IMAGE space at time t.
        
        Uses finite differences along the diffeomorphic path:
            v_teacher(t) = (x̃_{t+ε} - x̃_t) / ε
        
        Args:
            pod5_img: POD5 image [B, 1, 48, 48, 48] normalized to [-1, 1]
            poy1_img: POY1 image [B, 1, 48, 48, 48] normalized to [-1, 1]
            t: Time values [B] in [0, 1]
        
        Returns:
            v_teacher: Velocity in image space [B, 1, 48, 48, 48]
        """
        v_svf = self.get_svf(pod5_img, poy1_img)
        eps = self.fd_epsilon
        t_plus_eps = (t + eps).clamp(0.0, 1.0)
        x_t = self.get_warped_at_t(pod5_img, v_svf, t)
        x_t_eps = self.get_warped_at_t(pod5_img, v_svf, t_plus_eps)
        actual_eps = (t_plus_eps - t).view(-1, 1, 1, 1, 1).clamp(min=1e-6)
        v_teacher = (x_t_eps - x_t) / actual_eps
        return v_teacher
    
    @torch.no_grad()
    def get_full_path_velocity(
        self,
        pod5_img: torch.Tensor,
        poy1_img: torch.Tensor,
    ) -> torch.Tensor:
        """
        Get the full-path teacher velocity (t=0 to t=1) in image space.
        Returns the difference between the SVF-warped endpoint and POD5.
        """
        v_svf = self.get_svf(pod5_img, poy1_img)
        x1_warped = self.get_warped_at_t(
            pod5_img, v_svf,
            torch.ones(pod5_img.size(0), device=pod5_img.device)
        )
        return x1_warped - pod5_img
    
    @torch.no_grad()
    def validate_teacher_velocity(
        self,
        pod5_img: torch.Tensor,
        poy1_img: torch.Tensor,
        t: torch.Tensor,
        verbose: bool = True
    ) -> dict:
        """
        Validate teacher velocity computation in image space.
        """
        v_svf = self.get_svf(pod5_img, poy1_img)
        v_teacher = self.get_teacher_velocity_at_t(pod5_img, poy1_img, t)
        
        v_svf_mag = v_svf.abs().mean().item()
        v_teacher_mag = v_teacher.abs().mean().item()
        
        # Check SVF endpoint accuracy at t=1 (image space)
        x1_warped = self.get_warped_at_t(pod5_img, v_svf, torch.ones_like(t))
        svf_endpoint_error = (x1_warped - poy1_img).abs().mean().item()
        svf_endpoint_rel_error = svf_endpoint_error / (poy1_img.abs().mean().item() + 1e-8)
        
        # Velocity smoothness: compare at nearby t
        t_nearby = (t + 0.01).clamp(0.0, 1.0)
        v_nearby = self.get_teacher_velocity_at_t(pod5_img, poy1_img, t_nearby)
        velocity_smoothness = (v_teacher - v_nearby).abs().mean().item()
        
        # FD accuracy: compare at t=0 vs full-path velocity
        t_zero = torch.zeros_like(t)
        v_at_0 = self.get_teacher_velocity_at_t(pod5_img, poy1_img, t_zero)
        v_full = self.get_full_path_velocity(pod5_img, poy1_img)
        fd_accuracy = (v_at_0 - v_full).abs().mean().item()
        fd_rel_accuracy = fd_accuracy / (v_full.abs().mean().item() + 1e-8)
        
        metrics = {
            'v_svf_magnitude': v_svf_mag,
            'v_teacher_magnitude': v_teacher_mag,
            'svf_endpoint_error': svf_endpoint_error,
            'svf_endpoint_rel_error': svf_endpoint_rel_error,
            'velocity_smoothness': velocity_smoothness,
            'fd_accuracy_error': fd_accuracy,
            'fd_rel_accuracy': fd_rel_accuracy,
        }
        
        if verbose:
            print(f"\n{'='*70}")
            print(f"Teacher Velocity Validation (image space)")
            print(f"{'='*70}")
            print(f"1. Velocity Magnitudes:")
            print(f"   SVF velocity:          {v_svf_mag:.6f}")
            print(f"   Teacher velocity:      {v_teacher_mag:.6f}")
            print(f"\n2. SVF Endpoint Accuracy (t=1, image space):")
            print(f"   Absolute error: {svf_endpoint_error:.6f}")
            print(f"   Relative error: {svf_endpoint_rel_error:.4%}")
            print(f"   Status: {'✓ PASS' if svf_endpoint_rel_error < 0.20 else '✗ FAIL - SVF not reaching target!'}")
            print(f"\n3. Velocity Smoothness:")
            print(f"   Δv at nearby t: {velocity_smoothness:.6f}")
            print(f"   Status: {'✓ PASS' if velocity_smoothness < 0.1 else '⚠ WARNING - Velocity not smooth!'}")
            print(f"\n4. Finite Difference Accuracy (at t=0):")
            print(f"   FD vs full-path error: {fd_accuracy:.6f}")
            print(f"   Relative error: {fd_rel_accuracy:.4%}")
            print(f"   Status: {'✓ PASS' if fd_rel_accuracy < 0.20 else '⚠ WARNING - FD epsilon may be too large!'}")
            print(f"{'='*70}\n")
        
        return metrics



# ----------------------- Losses & Metrics -------------------------------
def bone_weight_map_from_poy1_hu(poy1_hu_t: torch.Tensor, threshold=BONE_HU_THRESHOLD, alpha=BONE_WEIGHT_ALPHA):
    """Build voxel weights from the target (POY1) HU.

    Used to emphasize bone regions in reconstruction losses:
      bone voxels (HU > threshold) -> weight = 1 + alpha
      other voxels                -> weight = 1
    """
    with torch.no_grad():
        bone = (poy1_hu_t > threshold).float()
        w = 1.0 + alpha * bone
    return w

def weighted_l1_norm(pred, target, weight):
    return (weight * torch.abs(pred - target)).sum() / (weight.sum() + 1e-8)

def tv_l2_3d(v: torch.Tensor) -> torch.Tensor:
    """ L2 TV on a vector field v [B,3,D,H,W]. """
    dz = v[:, :, 1:, :, :] - v[:, :, :-1, :, :]
    dy = v[:, :, :, 1:, :] - v[:, :, :, :-1, :]
    dx = v[:, :, :, :, 1:] - v[:, :, :, :, :-1]
    tv = (dz ** 2).mean() + (dy ** 2).mean() + (dx ** 2).mean()
    return tv

def ssim3d(x: torch.Tensor, y: torch.Tensor, C1=0.01**2, C2=0.03**2, win: int = 3) -> torch.Tensor:
    # Convert [-1,1] → [0,1] for proper constants
    x = (x + 1.0) * 0.5; y = (y + 1.0) * 0.5
    pad = win // 2
    pool = nn.AvgPool3d(win, stride=1, padding=0)
    xpad = F.pad(x, (pad,)*6, mode='replicate')
    ypad = F.pad(y, (pad,)*6, mode='replicate')
    mu_x = pool(xpad); mu_y = pool(ypad)
    mu_x2, mu_y2, mu_xy = mu_x**2, mu_y**2, mu_x*mu_y
    sigma_x2 = pool(xpad*xpad) - mu_x2
    sigma_y2 = pool(ypad*ypad) - mu_y2
    sigma_xy = pool(xpad*ypad) - mu_xy
    ssim_val = ((2*mu_xy + C1)*(2*sigma_xy + C2)) / ((mu_x2 + mu_y2 + C1)*(sigma_x2 + sigma_y2 + C2))
    return ssim_val.mean()

def ssim3d_map(x: torch.Tensor, y: torch.Tensor, data_range: float = 2.0, win: int = 3) -> torch.Tensor:
    """Return per-voxel SSIM map (same spatial shape as x/y) (FM V18 parity)."""
    C1 = (0.01 * float(data_range)) ** 2
    C2 = (0.03 * float(data_range)) ** 2
    pad = win // 2
    pool = nn.AvgPool3d(win, stride=1, padding=0)
    xpad = F.pad(x, (pad,) * 6, mode='replicate')
    ypad = F.pad(y, (pad,) * 6, mode='replicate')
    mu_x = pool(xpad)
    mu_y = pool(ypad)
    mu_x2, mu_y2, mu_xy = mu_x**2, mu_y**2, mu_x * mu_y
    sigma_x2 = pool(xpad * xpad) - mu_x2
    sigma_y2 = pool(ypad * ypad) - mu_y2
    sigma_xy = pool(xpad * ypad) - mu_xy
    ssim_val = ((2 * mu_xy + C1) * (2 * sigma_xy + C2)) / ((mu_x2 + mu_y2 + C1) * (sigma_x2 + sigma_y2 + C2))
    return ssim_val

def ms_ssim3d_simple(x: torch.Tensor, y: torch.Tensor, data_range: float = 2.0) -> torch.Tensor:
    """Simple MS-SSIM matching FM V18 exactly (expects normalized volumes, default data_range=2.0)."""
    C1 = (0.01 * float(data_range)) ** 2
    C2 = (0.03 * float(data_range)) ** 2

    def ssim3d_basic(a, b, C1=C1, C2=C2, win=3):
        pad = win // 2
        pool = nn.AvgPool3d(win, stride=1, padding=0)
        apad = F.pad(a, (pad,)*6, mode='replicate')
        bpad = F.pad(b, (pad,)*6, mode='replicate')
        mu_a = pool(apad)
        mu_b = pool(bpad)
        mu_a2, mu_b2, mu_ab = mu_a**2, mu_b**2, mu_a*mu_b
        sigma_a2 = pool(apad*apad) - mu_a2
        sigma_b2 = pool(bpad*bpad) - mu_b2
        sigma_ab = pool(apad*bpad) - mu_ab
        ssim = ((2*mu_ab + C1)*(2*sigma_ab + C2)) / ((mu_a2 + mu_b2 + C1)*(sigma_a2 + sigma_b2 + C2))
        return ssim.mean()
    s1 = ssim3d_basic(x, y)
    s2 = ssim3d_basic(F.avg_pool3d(x, 2), F.avg_pool3d(y, 2))
    return 0.5 * (s1 + s2)

def ms_ssim3d_masked(x: torch.Tensor, y: torch.Tensor, mask: torch.Tensor, eps: float = 1e-6) -> torch.Tensor:
    """Masked MS-SSIM: weight SSIM maps by a soft mask (e.g., bone voxels) across scales (FM V18 parity)."""
    if mask.dim() == 4:
        mask = mask.unsqueeze(1)
    mask = mask.to(dtype=x.dtype, device=x.device)
    # Scale 1
    s1_map = ssim3d_map(x, y, data_range=2.0)
    w1 = mask
    s1 = (s1_map * w1).sum() / (w1.sum() + eps)
    # Scale 2
    x2, y2 = F.avg_pool3d(x, 2), F.avg_pool3d(y, 2)
    w2 = F.avg_pool3d(w1, 2)
    s2_map = ssim3d_map(x2, y2, data_range=2.0)
    s2 = (s2_map * w2).sum() / (w2.sum() + eps)
    return 0.5 * (s1 + s2)

def compute_dice_score(pred_hu, target_hu, threshold=METRICS_BONE_HU_THRESHOLD) -> float:
    """FM V18 parity: Dice on HU-thresholded bone mask."""
    pred_bone = (pred_hu > threshold).astype(np.float32)
    target_bone = (target_hu > threshold).astype(np.float32)
    intersection = float(np.sum(pred_bone * target_bone))
    union_size = float(np.sum(pred_bone) + np.sum(target_bone))
    if union_size == 0.0:
        return 1.0
    return float((2.0 * intersection) / union_size)

def compute_comprehensive_metrics(pred_hu, target_hu, pred_norm=None, target_norm=None):
    """FM V18 parity metric set (5 metrics): MAE_all_HU, MAE_bone_HU, MS_SSIM, MS_SSIM_bone, Dice_bone."""
    metrics = {}
    metrics['MAE_all_HU'] = float(np.mean(np.abs(pred_hu - target_hu)))

    bone_mask = (target_hu > METRICS_BONE_HU_THRESHOLD)
    if bone_mask.any():
        metrics['MAE_bone_HU'] = float(np.mean(np.abs(pred_hu[bone_mask] - target_hu[bone_mask])))
    else:
        metrics['MAE_bone_HU'] = 0.0

    if pred_norm is not None and target_norm is not None:
        pred_t = torch.from_numpy(pred_norm).unsqueeze(0).unsqueeze(0).float().to(device)
        target_t = torch.from_numpy(target_norm).unsqueeze(0).unsqueeze(0).float().to(device)
    else:
        pred_clip = np.clip(pred_hu, HU_RANGE[0], HU_RANGE[1])
        target_clip = np.clip(target_hu, HU_RANGE[0], HU_RANGE[1])
        pred_norm_tmp = 2.0 * (pred_clip - HU_RANGE[0]) / (HU_RANGE[1] - HU_RANGE[0]) - 1.0
        target_norm_tmp = 2.0 * (target_clip - HU_RANGE[0]) / (HU_RANGE[1] - HU_RANGE[0]) - 1.0
        pred_t = torch.from_numpy(pred_norm_tmp).unsqueeze(0).unsqueeze(0).float().to(device)
        target_t = torch.from_numpy(target_norm_tmp).unsqueeze(0).unsqueeze(0).float().to(device)

    with torch.no_grad():
        metrics['MS_SSIM'] = float(ms_ssim3d_simple(pred_t, target_t).item())
        if bone_mask.any():
            bone_mask_t = torch.from_numpy(bone_mask.astype(np.float32)).unsqueeze(0).unsqueeze(0).to(device)
            metrics['MS_SSIM_bone'] = float(ms_ssim3d_masked(pred_t, target_t, bone_mask_t).item())
        else:
            metrics['MS_SSIM_bone'] = 0.0

    metrics['Dice_bone'] = compute_dice_score(pred_hu, target_hu, threshold=METRICS_BONE_HU_THRESHOLD)
    return metrics

def _make_w_slab_mask_np(shape: tuple[int, int, int], w_start: int, w_end: int) -> np.ndarray:
    """FM V18 parity: hard slab mask selecting W slices in [w_start, w_end] inclusive."""
    if len(shape) != 3:
        raise ValueError(f"Expected 3D shape (D,H,W), got: {shape}")
    D, H, W = shape
    if W <= 0:
        return np.zeros((D, H, W), dtype=bool)
    w_start_i = int(w_start)
    w_end_i = int(w_end)
    if w_end_i < w_start_i:
        w_start_i, w_end_i = w_end_i, w_start_i
    w_start_i = max(0, min(W - 1, w_start_i))
    w_end_i = max(0, min(W - 1, w_end_i))
    mask = np.zeros((D, H, W), dtype=bool)
    mask[:, :, w_start_i:w_end_i + 1] = True
    return mask

def compute_comprehensive_metrics_middle_slab(pred_hu, target_hu, pred_norm=None, target_norm=None):
    """FM V18 parity: same 5 metrics but restricted to middle-slab W slices."""
    w_start = int(globals().get('MIDDLE_SLAB_IMAGE_SLICE_START', 20))
    w_end = int(globals().get('MIDDLE_SLAB_IMAGE_SLICE_END', 28))
    slab_mask = _make_w_slab_mask_np(target_hu.shape, w_start, w_end)
    metrics = {}

    if slab_mask.any():
        metrics['MAE_all_HU_mid'] = float(np.mean(np.abs(pred_hu[slab_mask] - target_hu[slab_mask])))
    else:
        metrics['MAE_all_HU_mid'] = float('nan')

    bone_mask = (target_hu > METRICS_BONE_HU_THRESHOLD) & slab_mask
    if bone_mask.any():
        metrics['MAE_bone_HU_mid'] = float(np.mean(np.abs(pred_hu[bone_mask] - target_hu[bone_mask])))
    else:
        metrics['MAE_bone_HU_mid'] = 0.0

    if pred_norm is not None and target_norm is not None:
        pred_t = torch.from_numpy(pred_norm).unsqueeze(0).unsqueeze(0).float().to(device)
        target_t = torch.from_numpy(target_norm).unsqueeze(0).unsqueeze(0).float().to(device)
    else:
        pred_clip = np.clip(pred_hu, HU_RANGE[0], HU_RANGE[1])
        target_clip = np.clip(target_hu, HU_RANGE[0], HU_RANGE[1])
        pred_norm_tmp = 2.0 * (pred_clip - HU_RANGE[0]) / (HU_RANGE[1] - HU_RANGE[0]) - 1.0
        target_norm_tmp = 2.0 * (target_clip - HU_RANGE[0]) / (HU_RANGE[1] - HU_RANGE[0]) - 1.0
        pred_t = torch.from_numpy(pred_norm_tmp).unsqueeze(0).unsqueeze(0).float().to(device)
        target_t = torch.from_numpy(target_norm_tmp).unsqueeze(0).unsqueeze(0).float().to(device)

    slab_mask_t = torch.from_numpy(slab_mask.astype(np.float32)).unsqueeze(0).unsqueeze(0).to(device)
    with torch.no_grad():
        metrics['MS_SSIM_mid'] = float(ms_ssim3d_masked(pred_t, target_t, slab_mask_t).item()) if slab_mask.any() else float('nan')
        if bone_mask.any():
            bone_mask_t = torch.from_numpy(bone_mask.astype(np.float32)).unsqueeze(0).unsqueeze(0).to(device)
            metrics['MS_SSIM_bone_mid'] = float(ms_ssim3d_masked(pred_t, target_t, bone_mask_t).item())
        else:
            metrics['MS_SSIM_bone_mid'] = 0.0

    pred_bin = (pred_hu > METRICS_BONE_HU_THRESHOLD) & slab_mask
    tgt_bin = (target_hu > METRICS_BONE_HU_THRESHOLD) & slab_mask
    inter = float(np.logical_and(pred_bin, tgt_bin).sum())
    denom = float(pred_bin.sum() + tgt_bin.sum())
    metrics['Dice_bone_mid'] = float((2.0 * inter) / (denom + 1e-6)) if denom > 0 else 0.0
    return metrics


# ----------------------- Excel Helper -----------------------------
def write_metrics_excel(df: pd.DataFrame, path: Path):
    """Write metrics DataFrame to Excel with wider cells and 3-decimal numeric format."""
    path.parent.mkdir(parents=True, exist_ok=True)
    excel_path = str(path)
    try:
        import xlsxwriter  # type: ignore  # noqa: F401
        with pd.ExcelWriter(excel_path, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="metrics")
            wb = writer.book
            ws = writer.sheets["metrics"]
            header_fmt = wb.add_format({"bold": True})
            num_fmt = wb.add_format({"num_format": "0.000"})
            text_fmt = wb.add_format()
            ws.set_row(0, 22, header_fmt)
            for col_idx, col_name in enumerate(df.columns):
                series = df[col_name]
                try:
                    max_data = int(series.astype(str).map(len).max()) if len(series) > 0 else 0
                except Exception:
                    max_data = 0
                width = min(max(max(len(str(col_name)), max_data) + 2, 12), 32)
                if pd.api.types.is_numeric_dtype(series):
                    ws.set_column(col_idx, col_idx, width, num_fmt)
                else:
                    ws.set_column(col_idx, col_idx, width, text_fmt)
            for r in range(1, len(df) + 1):
                ws.set_row(r, 18)
        return
    except Exception:
        pass
    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="metrics")
            wb = writer.book
            ws = writer.sheets["metrics"]
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.row_dimensions[1].height = 22
            for col_idx, col_name in enumerate(df.columns, start=1):
                series = df[col_name]
                try:
                    max_data = int(series.astype(str).map(len).max()) if len(series) > 0 else 0
                except Exception:
                    max_data = 0
                width = min(max(max(len(str(col_name)), max_data) + 2, 12), 32)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
                if pd.api.types.is_numeric_dtype(series):
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
                        for cell in row:
                            cell.number_format = '0.000'
            for r in range(2, ws.max_row + 1):
                ws.row_dimensions[r].height = 18
        return
    except Exception:
        pass
    df.to_excel(excel_path, index=False)

def create_metrics_excel_with_footnotes(metrics_list, save_path):
    """FM V18 parity: write Epoch Metrics + Metric Definitions, with fixed column order."""
    df = pd.DataFrame(metrics_list)

    col_order = ['epoch']
    train_cols = ['avg_total_loss', 'avg_fm_loss', 'avg_endpoint_loss']
    for c in train_cols:
        if c in df.columns:
            col_order.append(c)

    test_cols = [
        'MAE_all_HU', 'MAE_bone_HU', 'MS_SSIM', 'MS_SSIM_bone', 'Dice_bone',
        'MAE_all_HU_mid', 'MAE_bone_HU_mid', 'MS_SSIM_mid', 'MS_SSIM_bone_mid', 'Dice_bone_mid',
        'Jacobian_min', 'Jacobian_mean', 'Jacobian_nonpos_frac', 'Diffeomorphic'
    ]
    missing_mid = [
        c for c in ['MAE_all_HU_mid', 'MAE_bone_HU_mid', 'MS_SSIM_mid', 'MS_SSIM_bone_mid', 'Dice_bone_mid']
        if c not in df.columns
    ]
    if missing_mid:
        print(f"⚠️ Middle-slab metric columns missing (won't appear in Excel): {missing_mid}")
    for c in test_cols:
        if c in df.columns:
            col_order.append(c)

    remaining = [c for c in df.columns if c not in col_order]
    col_order.extend(remaining)
    df = df[col_order]

    avg_row = df.select_dtypes(include=[np.number]).mean()
    avg_row['epoch'] = 'AVERAGE'
    std_row = df.select_dtypes(include=[np.number]).std()
    std_row['epoch'] = 'STD_DEV'
    df = pd.concat([df, pd.DataFrame([avg_row]), pd.DataFrame([std_row])], ignore_index=True)

    num_cols = df.select_dtypes(include=[np.number]).columns
    df[num_cols] = df[num_cols].round(4)

    try:
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Epoch Metrics')
            notes = pd.DataFrame({
                'Metric': [
                    'MAE_all_HU', 'MAE_bone_HU', 'MS_SSIM', 'MS_SSIM_bone', 'Dice_bone',
                    'MAE_all_HU_mid', 'MAE_bone_HU_mid', 'MS_SSIM_mid', 'MS_SSIM_bone_mid', 'Dice_bone_mid',
                    'Jacobian_min', 'Jacobian_mean', 'Jacobian_nonpos_frac', 'Diffeomorphic'
                ],
                'Meaning': [
                    'Mean absolute error in HU (all voxels)',
                    'Mean absolute error in HU (bone voxels only, >threshold)',
                    'Multi-scale SSIM (similarity, 0-1, higher is better)',
                    'Multi-scale SSIM computed on bone region only',
                    'Dice coefficient for bone segmentation',
                    'MAE_all_HU computed only on the middle slab W slices',
                    'MAE_bone_HU computed only on the middle slab W slices',
                    'MS_SSIM computed only on the middle slab W slices',
                    'MS_SSIM_bone computed only on the middle slab W slices',
                    'Dice_bone computed only on the middle slab W slices',
                    'Minimum Jacobian determinant (should be >0 for diffeomorphic warp)',
                    'Mean Jacobian determinant (≈1.0 for volume-preserving transform)',
                    'Fraction of voxels with non-positive Jacobian (should be 0 for diffeomorphic)',
                    'True if all Jacobian determinants are positive (fully diffeomorphic)'
                ]
            })
            notes.to_excel(writer, index=False, sheet_name='Metric Definitions')

            ws = writer.book["Epoch Metrics"]
            for col_cells in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max(10, max_len + 2), 35)
    except Exception as e:
        print(f"⚠️ Failed to write FM V18-style Excel (openpyxl). Falling back to basic Excel: {e}")
        df.to_excel(str(save_path), index=False)

    print(f"📊 Metrics saved to: {save_path}")

def dice_coefficient(m1: torch.Tensor, m2: torch.Tensor, eps: float = 1e-6) -> torch.Tensor:
    """DICE coefficient matching FM V10: 2*|A∩B| / (|A| + |B|)."""
    inter = (m1 * m2).sum()
    union_size = m1.sum() + m2.sum()
    if union_size == 0:
        return torch.tensor(1.0, device=m1.device)  # Both empty, perfect match
    return (2.0 * inter) / union_size

# ==================== Semi-Online Augmentation Sampler ====================
class SemiOnlineAugmentationSampler(torch.utils.data.Sampler):
    """
    Sampler for semi-online augmentation:
    - For each ROI, picks aug0 + NUM_AUG_PER_ROI random augmentations each epoch
    - Different random selection each epoch
    - Works with Subset objects from train_test_split
    """
    def __init__(self, dataset, num_random_aug_per_roi: int = NUM_AUG_PER_ROI, seed: int = RANDOM_SEED):
        self.dataset = dataset
        self.num_random_aug_per_roi = max(0, int(num_random_aug_per_roi))
        self.rng = np.random.RandomState(seed)
        
        # Handle Subset vs raw Dataset
        from torch.utils.data import Subset
        self.is_subset = isinstance(dataset, Subset)
        if self.is_subset:
            self.base_dataset = dataset.dataset
            self.valid_indices = set(dataset.indices)
            self.base_to_subset_idx = {base_idx: subset_idx for subset_idx, base_idx in enumerate(dataset.indices)}
        else:
            self.base_dataset = dataset
            self.valid_indices = set(range(len(dataset)))
            self.base_to_subset_idx = None
        
        self._build_group_indices()

    def _build_group_indices(self):
        """Build grouping of indices by (case_id, roi_num), pre-separated by aug_id."""
        self.group_indices = {}
        for idx in self.valid_indices:
            pair = self.base_dataset.pairs[idx]
            key = (pair["case_id"], pair["roi_num"])
            if key not in self.group_indices:
                self.group_indices[key] = {"aug0": [], "others": []}
            
            if pair["aug_id"] == 0:
                self.group_indices[key]["aug0"].append(idx)
            else:
                self.group_indices[key]["others"].append(idx)

    def __len__(self):
        total = 0
        for key, group in self.group_indices.items():
            aug0_count = len(group["aug0"])
            other_count = len(group["others"])
            if aug0_count == 0:
                k = min(aug0_count + other_count, 1 + self.num_random_aug_per_roi)
                total += k
            else:
                k = min(other_count, self.num_random_aug_per_roi)
                total += 1 + k
        return total

    def __iter__(self):
        all_indices = []
        for key, group in self.group_indices.items():
            orig_indices = group["aug0"]
            other_indices = group["others"]
            
            # Always include first aug0 if available
            if orig_indices:
                all_indices.append(orig_indices[0])
            
            # Randomly sample from other augmentations
            if other_indices and self.num_random_aug_per_roi > 0:
                n_pick = min(len(other_indices), self.num_random_aug_per_roi)
                picked = self.rng.choice(other_indices, size=n_pick, replace=False)
                all_indices.extend(picked.tolist())
            
            # Edge case fallback
            elif not orig_indices and other_indices:
                n_pick = min(len(other_indices), 1 + self.num_random_aug_per_roi)
                picked = self.rng.choice(other_indices, size=n_pick, replace=False)
                all_indices.extend(picked.tolist())
        
        # Shuffle and convert to subset indices if needed
        self.rng.shuffle(all_indices)
        if self.is_subset:
            for base_idx in all_indices:
                yield self.base_to_subset_idx[base_idx]
        else:
            for idx in all_indices:
                yield idx

# ----------------------- Dataset Splitting -------------------
def split_dataset(full_dataset, train_split=TRAIN_SPLIT, seed=RANDOM_SEED, split_by_patient=SPLIT_BY_PATIENT,
                  use_semi_online=USE_SEMI_ONLINE_AUG, max_aug_id=NUM_AUG_PER_ROI):
    """
    Split dataset BY PATIENT or BY ROI to prevent data leakage from augmentations.
    
    Strategy:
    - If split_by_patient=True: Split by patient ID (all ROIs of same patient stay together)
    - If split_by_patient=False: Split by individual ROI (each ROI is independent)
    - Train set: Uses augmentations based on use_semi_online flag
      - If use_semi_online=True: ALL augmentations available (sampler picks randomly each epoch)
      - If use_semi_online=False: Only aug0 through max_aug_id (deterministic subset)
    - Test set: ONLY original (aug0) - no augmentations
    
    Args:
        full_dataset: Complete ROI3DDataset
        train_split: Fraction of units for training (0.85 = 85%)
        seed: Random seed for reproducibility
        split_by_patient: True=split by patient ID, False=split by ROI independently
        use_semi_online: If True, allow all augmentations (sampler controls usage); if False, filter to max_aug_id
        max_aug_id: Maximum augmentation ID to include when use_semi_online=False
    
    Returns:
        train_dataset, test_dataset (as Subset objects)
    """
    from torch.utils.data import Subset
    
    if split_by_patient:
        unique_keys = sorted(set(p["case_id"] for p in full_dataset.pairs))
        unit_name = "patient"
        get_unit_id = lambda p: p["case_id"]
    else:
        unique_keys = sorted(set((p["case_id"], p["roi_num"]) for p in full_dataset.pairs))
        unit_name = "ROI"
        get_unit_id = lambda p: (p["case_id"], p["roi_num"])
    
    n_keys = len(unique_keys)
    
    # Use RandomState.shuffle() to match FM V10/V15 and VAE V10 split logic exactly
    rng = np.random.RandomState(seed)
    key_indices = np.arange(n_keys)
    rng.shuffle(key_indices)
    n_train_keys = int(train_split * n_keys)
    train_keys = set(unique_keys[i] for i in key_indices[:n_train_keys])
    test_keys = set(unique_keys[i] for i in key_indices[n_train_keys:])
    
    mode_str = f"semi-online (all augs available)" if use_semi_online else f"fixed (aug0-aug{max_aug_id})"
    print(f"📊 {unit_name.capitalize()}-Based Dataset Split ({mode_str}):")
    print(f"   Total unique {unit_name}s: {n_keys}")
    print(f"   Train {unit_name}s: {len(train_keys)}")
    print(f"   Test {unit_name}s: {len(test_keys)}")
    
    # Build train indices based on augmentation mode
    train_idx = []
    for idx, pair in enumerate(full_dataset.pairs):
        if get_unit_id(pair) in train_keys:
            aug_id = pair.get("aug_id", 0)
            # Filter augmentations based on mode
            if not use_semi_online and aug_id > max_aug_id:
                continue  # Skip augmentations beyond max_aug_id when not using semi-online
            train_idx.append(idx)
    
    # Build test indices: ONLY original (aug0) for test units
    test_idx = []
    for idx, pair in enumerate(full_dataset.pairs):
        if get_unit_id(pair) in test_keys:
            # Only include original (aug_id == 0)
            if pair["aug_id"] == 0:
                test_idx.append(idx)
    
    print(f"   Train samples: {len(train_idx)} ({'all augs' if use_semi_online else f'aug0-aug{max_aug_id}'})")
    print(f"   Test samples: {len(test_idx)} (original only)")
    
    return Subset(full_dataset, train_idx), Subset(full_dataset, test_idx)


@torch.no_grad()
def predict(model, x0_img, x1_img, use_cfg=False, cfg_weight=1.5):
    """
    Run one-shot diffeomorphic registration via SVF (VELOCITY ONLY, no intensity).
    
    Mathematical Model:
        - v_svf = model(x0, x1): predict stationary velocity field (3 channels)
        - φ = exp(v_svf): diffeomorphic warp via scaling-and-squaring
        - x̂_1 = warp(x0, φ): pure diffeomorphic warping (no intensity residual)
      
    Args:
        x0_img: Moving image (POD5) in IMAGE space [B, 1, 48, 48, 48]
        x1_img: Fixed image (POY1) in IMAGE space [B, 1, 48, 48, 48]
        use_cfg: If True, generate both unconditional (x0-only) and conditional (x0+x1) predictions
        cfg_weight: Guidance weight for CFG: final = uncond + w * (cond - uncond)
    
    Returns:
        If use_cfg=False: (x_hat1, v_vox, phi_norm)
        If use_cfg=True: (x_hat1_guided, v_vox, phi_norm, x_hat1_uncond, x_hat1_cond)
        Note: x_hat1 is always in IMAGE space for metrics/visualization
    """
    x0_input = x0_img
    x0 = x0_input
    x1 = x1_img

    if use_cfg and not model.training:
        # CFG inference: generate both unconditional and conditional predictions
        flag_shape = x0[:, :1, ...]
        flag_uncond = torch.zeros_like(flag_shape)
        flag_cond = torch.ones_like(flag_shape)
        v_vox_uncond = model(x0, x0, cond_flag=flag_uncond, t=None)  # VELOCITY ONLY
        v_vox_cond = model(x0, x1, cond_flag=flag_cond, t=None)      # VELOCITY ONLY
        
        # Apply resection plane constraint to velocities
        if RESECTION_PLANE_CONSTRAINT:
            v_vox_uncond = apply_resection_plane_mask(v_vox_uncond, sigma=RESECTION_PLANE_SIGMA)
            v_vox_cond = apply_resection_plane_mask(v_vox_cond, sigma=RESECTION_PLANE_SIGMA)
        
        # SVF mode: compute phi and warp (mask already applied above)
        phi_norm_uncond = expv_scaling_squaring(v_vox_uncond, apply_resection_mask=False)
        phi_norm_cond = expv_scaling_squaring(v_vox_cond, apply_resection_mask=False)
        x_hat1_uncond_warp = warp_image_with_phi_norm(x0, phi_norm_uncond)  # Pure warping
        x_hat1_cond_warp = warp_image_with_phi_norm(x0, phi_norm_cond)      # Pure warping
        
        # Apply CFG in image space
        x_hat1_guided = (x_hat1_uncond_warp + cfg_weight * (x_hat1_cond_warp - x_hat1_uncond_warp)).clamp(-1, 1)
        x_hat1_uncond = x_hat1_uncond_warp.clamp(-1, 1)
        x_hat1_cond = x_hat1_cond_warp.clamp(-1, 1)
        
        return x_hat1_guided, v_vox_cond, phi_norm_cond, x_hat1_uncond, x_hat1_cond
    
    else:
        # Standard inference: single prediction using SVF (VELOCITY ONLY)
        flag_shape = x0[:, :1, ...]
        flag_cond = torch.ones_like(flag_shape) if globals().get('USE_CFG_DROPOUT', False) else None
        v_vox = model(x0, x1, cond_flag=flag_cond, t=None)  # VELOCITY ONLY (3 channels)
        
        # Apply resection plane constraint to velocity
        if RESECTION_PLANE_CONSTRAINT:
            v_vox = apply_resection_plane_mask(v_vox, sigma=RESECTION_PLANE_SIGMA)
        
        # SVF mode: compute phi via scaling-and-squaring and warp
        phi_norm = expv_scaling_squaring(v_vox, apply_resection_mask=False)
        x_hat1 = warp_image_with_phi_norm(x0, phi_norm).clamp(-1, 1)
        
        return x_hat1, v_vox, phi_norm





@torch.no_grad()
def evaluate_model(model, data_loader):
    model.eval()
    per_sample_metrics = []
    all_det_min, all_det_mean, all_det_nonpos_frac = [], [], []
    n_batches = len(data_loader)
    pbar = tqdm(total=n_batches, desc="Eval", leave=False) if HAS_TQDM else None
    for batch in data_loader:
        x0 = batch["x0"].to(device)
        x1 = batch["x1"].to(device)
        x_hat1, v_vox, phi_norm = predict(model, x0, x1)  # VELOCITY ONLY (3 returns)

        # Jacobian determinant metrics (diffeomorphism check) - only for SVF mode
        if phi_norm is not None:
            detJ = jacobian_determinant(phi_norm).detach()
            det_min = float(detJ.min().cpu().item())
            det_mean = float(detJ.mean().cpu().item())
            nonpos_frac = float((detJ <= 0).float().mean().cpu().item())
            all_det_min.append(det_min)
            all_det_mean.append(det_mean)
            all_det_nonpos_frac.append(nonpos_frac)

        # FM V18 parity: compute metrics per-sample (includes middle-slab variants)
        pred_hu_b = denorm_to_hu(x_hat1.squeeze(1).detach().cpu().numpy(), HU_RANGE)
        gt_src = x1
        gt_hu_b = denorm_to_hu(gt_src.squeeze(1).detach().cpu().numpy(), HU_RANGE)
        pred_norm_b = x_hat1.squeeze(1).detach().cpu().numpy()
        gt_norm_b = gt_src.squeeze(1).detach().cpu().numpy()

        B = pred_hu_b.shape[0]
        for i in range(B):
            m_full = compute_comprehensive_metrics(
                pred_hu=pred_hu_b[i],
                target_hu=gt_hu_b[i],
                pred_norm=pred_norm_b[i],
                target_norm=gt_norm_b[i],
            )
            m_mid = compute_comprehensive_metrics_middle_slab(
                pred_hu=pred_hu_b[i],
                target_hu=gt_hu_b[i],
                pred_norm=pred_norm_b[i],
                target_norm=gt_norm_b[i],
            )
            per_sample_metrics.append({**m_full, **m_mid})

        if pbar: pbar.update(1)
    if pbar: pbar.close()
    
    # Jacobian metrics only available for SVF mode (not FM mode where phi_norm=None)
    if all_det_min:
        jac_metrics = {
            'Jacobian_min': float(np.mean(all_det_min)),
            'Jacobian_mean': float(np.mean(all_det_mean)),
            'Jacobian_nonpos_frac': float(np.mean(all_det_nonpos_frac)),
            'Diffeomorphic': bool(np.mean(all_det_nonpos_frac) == 0.0),
        }
    else:
        # FM mode: no Jacobian metrics (no displacement field)
        jac_metrics = {
            'Jacobian_min': float('nan'),
            'Jacobian_mean': float('nan'),
            'Jacobian_nonpos_frac': float('nan'),
            'Diffeomorphic': None,
        }
    
    # Aggregate metrics (FM V18-style keys)
    if per_sample_metrics:
        dfm = pd.DataFrame(per_sample_metrics)
        agg = dfm.mean(numeric_only=True).to_dict()
    else:
        agg = {
            'MAE_all_HU': float('nan'),
            'MAE_bone_HU': float('nan'),
            'MS_SSIM': float('nan'),
            'MS_SSIM_bone': float('nan'),
            'Dice_bone': float('nan'),
            'MAE_all_HU_mid': float('nan'),
            'MAE_bone_HU_mid': float('nan'),
            'MS_SSIM_mid': float('nan'),
            'MS_SSIM_bone_mid': float('nan'),
            'Dice_bone_mid': float('nan'),
        }

    metrics = {**agg, **jac_metrics}
    return metrics

def train():
    dataset = ROI3DDataset(POD5_DIR, POY1_DIR, normalize=True)
    train_ds, test_ds = split_dataset(dataset, train_split=TRAIN_SPLIT, seed=RANDOM_SEED, split_by_patient=SPLIT_BY_PATIENT)

    # Create dataloaders with optional semi-online augmentation
    if USE_SEMI_ONLINE_AUG:
        print(f"🔁 Semi-online augmentation enabled: original + {NUM_AUG_PER_ROI} random augs per ROI each epoch")
        train_sampler = SemiOnlineAugmentationSampler(train_ds, num_random_aug_per_roi=NUM_AUG_PER_ROI, seed=RANDOM_SEED)
        train_loader = DataLoader(train_ds, batch_size=BATCH_SIZE, sampler=train_sampler, num_workers=0)
        actual_train_samples_per_epoch = len(train_sampler)
        print(f"📊 Training: {len(train_ds)} available samples → {actual_train_samples_per_epoch} used per epoch (~{actual_train_samples_per_epoch // BATCH_SIZE} batches)")
    else:
        train_loader = DataLoader(train_ds, batch_size=BATCH_SIZE, shuffle=True, num_workers=0)
        print(f"📊 Training: {len(train_ds)} samples per epoch (~{len(train_ds) // BATCH_SIZE} batches)")
    
    test_loader = DataLoader(test_ds, batch_size=BATCH_SIZE, shuffle=False, num_workers=0)
    print(f"📊 Test: {len(test_ds)} samples (~{len(test_ds) // BATCH_SIZE} batches)")
    print()

    base_ch = UNET_BASE_CHANNELS
    # SVF mode: no time embedding needed (we use scaling-and-squaring instead of ODE integration)
    model = UNet3D(base_channels=base_ch, use_time_emb=False).to(device)
    optimizer = torch.optim.AdamW(model.parameters(), lr=LEARNING_RATE, weight_decay=1e-5)


    print(f"Training samples: {len(train_ds)}, Test samples: {len(test_ds)}")
    print(f"Model Params: {sum(p.numel() for p in model.parameters()):,}")
    print(f"\n📋 Configuration:")
    print(f"   Deformation mode: SVF (Diffeomorphic via Stationary Velocity Field)")
    
    print(f"   Space mode: IMAGE (48³)")
    print(f"   Input: [B, 2, 48, 48, 48] (POD5 + POY1 concatenated)")
    print(f"   Output: [B, 3, 48, 48, 48] (VELOCITY ONLY: vz, vy, vx)")
    print(f"   Mathematical model: φ = exp(v_svf) via scaling-and-squaring (7 iterations)")
    print(f"   Loss: L1 + MS-SSIM + TV(v_vox) [NO intensity penalty]")
    print(f"   Base channels: {base_ch}")
    print()
    
    # Resection plane constraints
    if RESECTION_PLANE_CONSTRAINT:
        spatial_size = ROI_SHAPE[2]  # W dimension for image space
        info = get_resection_plane_weight_info(spatial_size, RESECTION_PLANE_SIGMA, profile=RESECTION_PLANE_PROFILE)
        print(f"🔒 Deformation constraint: ENABLED (masking velocity)")
        print(f"   profile={info['profile']}, σ={info['sigma']:.1f}, center W={info['center']:.1f}, min_w={info['min_weight']}")
        print(f"   w@d=0.5: {info['w_d0p5_left']:.6f} (both middle slices), w@d=1.5: {info['w_d1p5']:.6f}, w@d=2.5: {info['w_d2p5']:.6f}")
        print(f"   w_edge={info['w_edge']:.6f}, max_w_in_ROI={info['w_max']:.6f}")
    if INTENSITY_PLANE_CONSTRAINT:
        spatial_size = ROI_SHAPE[2]
        info_int = get_resection_plane_weight_info(spatial_size, INTENSITY_PLANE_SIGMA, profile=RESECTION_PLANE_PROFILE)
        print(f"🔒 Intensity constraint: ENABLED (masking intensity residual)")
        print(f"   profile={info_int['profile']}, σ={info_int['sigma']:.1f}, center W={info_int['center']:.1f}")
        print(f"   w@d=0.5: {info_int['w_d0p5_left']:.6f}, w@d=1.5: {info_int['w_d1p5']:.6f}, w@d=2.5: {info_int['w_d2p5']:.6f}")
        print(f"   w_edge={info_int['w_edge']:.6f}, max_w_in_ROI={info_int['w_max']:.6f}")
    if RESECTION_PLANE_CONSTRAINT or INTENSITY_PLANE_CONSTRAINT:
        print(f"   ROI axes: D=U, H=V, W=N (normal to plane)")

    metrics_rows = []
    if HAS_TQDM:
        epoch_pbar = tqdm(total=NUM_EPOCHS, desc="Training Progress", position=0)

    for epoch in range(1, NUM_EPOCHS+1):
        model.train()
        losses = []
        n_batches = len(train_loader)
        batch_pbar = tqdm(total=n_batches, desc=f"Epoch {epoch}/{NUM_EPOCHS}", leave=False, position=1) if HAS_TQDM else None

        for batch in train_loader:
            x0_img = batch["x0"].to(device)
            x1_img = batch["x1"].to(device)
            x1_hu = batch["x1_hu"].to(device)
            x0 = x0_img
            x1 = x1_img

            # CFG-style dropout: randomly drop POY1 (fixed) image during training
            if USE_CFG_DROPOUT and model.training:
                # Per-sample dropout mask: 1 = keep x1, 0 = replace
                B = x0.size(0)
                keep_mask = (torch.rand(B, device=device) >= CFG_DROPOUT_PROB).float()
                keep_mask = keep_mask.view(B, 1, 1, 1, 1)  # [B,1,1,1,1] for broadcasting
                
                # Choose replacement strategy for dropped samples
                if CFG_DROPOUT_REPLACE == 'zeros':
                    # Original CFG: replace with zeros (can cause distribution shift)
                    x1_replacement = torch.zeros_like(x1)
                elif CFG_DROPOUT_REPLACE == 'noise':
                    # Replace with Gaussian noise matching x1 statistics (recommended)
                    # Sample noise in same range as normalized images [-1, 1]
                    x1_replacement = torch.randn_like(x1) * 0.5  # std=0.5 for [-1,1] range
                elif CFG_DROPOUT_REPLACE == 'copy_x0':
                    # Replace with POD5 (no future information, forces unconditional learning)
                    x1_replacement = x0
                else:
                    raise ValueError(f"Unknown CFG_DROPOUT_REPLACE: {CFG_DROPOUT_REPLACE}")
                
                # Apply mask: keep x1 where mask=1, use replacement where mask=0
                x1_input = x1 * keep_mask + x1_replacement * (1.0 - keep_mask)
            else:
                x1_input = x1

            # =========================================================================
            # SVF Diffeomorphic Registration Training
            # =========================================================================
            # Mathematical Foundation:
            #   v_svf(ξ): stationary velocity field over spatial coordinates ξ∈Ω
            #   φ = exp(v_svf): diffeomorphic warp via scaling-and-squaring (7 iterations)
            #   x̂_1 = warp(x0, φ) + a: warp moving image and add intensity residual
            #
            # This produces biologically-plausible diffeomorphic deformations that
            # respect smooth, topology-preserving transformations.
            # =========================================================================
            
            # Provide conditioning flag: 1 when x1_input differs from x0, else 0
            cond_flag = (x1_input != x0).float().mean(dim=1, keepdim=True)
            v_vox = model(x0, x1_input, cond_flag=cond_flag, t=None)  # VELOCITY ONLY (no intensity)
            
            # Apply resection plane constraint to velocity BEFORE computing phi
            # This ensures the constraint affects both warping AND regularization losses
            if RESECTION_PLANE_CONSTRAINT:
                v_vox = apply_resection_plane_mask(v_vox, sigma=RESECTION_PLANE_SIGMA)
            
            # Exponentiate SVF to diffeomorphic warp: φ = exp(v_svf)
            phi_norm = expv_scaling_squaring(v_vox, apply_resection_mask=False)  # Already masked above
            
            # VELOCITY-ONLY: No intensity residual, just warp
            # x̂_1 = x0 ∘ φ (pure diffeomorphic warping, no additive component)
            x_warp = warp_image_with_phi_norm(x0, phi_norm)
            x_hat1 = x_warp  # No intensity residual added
            
            # IMAGE SPACE LOSS: bone-weighted reconstruction
            x_hat1 = x_hat1.clamp(-1, 1)
            w_bone = bone_weight_map_from_poy1_hu(x1_hu)
            l1_loss = weighted_l1_norm(x_hat1, x1, w_bone)
            ms_ssim = ms_ssim3d_simple(x_hat1, x1)
            recon_loss = 0.1 * (1.0 - ms_ssim) + 0.05 * l1_loss
            tv_loss = tv_l2_3d(v_vox)
            tv_w = 2e-2
            loss = recon_loss + tv_w * tv_loss
            
            if not loss.requires_grad:
                # Debug: trace where gradient breaks
                import sys
                print(f"DEBUG: v_vox.requires_grad = {v_vox.requires_grad}", flush=True)
                print(f"DEBUG: phi_norm.requires_grad = {phi_norm.requires_grad}", flush=True)
                print(f"DEBUG: x_warp.requires_grad = {x_warp.requires_grad}", flush=True)
                print(f"DEBUG: x_hat1.requires_grad = {x_hat1.requires_grad}", flush=True)
                print(f"DEBUG: tv_loss.requires_grad = {tv_loss.requires_grad}", flush=True)
                sys.stdout.flush()
                raise RuntimeError("Deformable loss has no grad; check graph inputs.")
            
            optimizer.zero_grad(set_to_none=True)
            loss.backward()
            torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
            optimizer.step()

            losses.append(float(loss.item()))
            if batch_pbar:
                batch_pbar.set_postfix({'loss': f"{loss.item():.4f}"})
                batch_pbar.update(1)

        if batch_pbar: batch_pbar.close()

        # Eval
        model.eval()
        test_metrics = evaluate_model(model, test_loader)
        row = {
            "epoch": epoch,
            "avg_total_loss": float(np.mean(losses)) if losses else float('nan'),
            "MAE_all_HU": test_metrics.get('MAE_all_HU', float('nan')),
            "MAE_bone_HU": test_metrics.get('MAE_bone_HU', float('nan')),
            "MS_SSIM": test_metrics.get('MS_SSIM', float('nan')),
            "MS_SSIM_bone": test_metrics.get('MS_SSIM_bone', float('nan')),
            "Dice_bone": test_metrics.get('Dice_bone', float('nan')),
            "MAE_all_HU_mid": test_metrics.get('MAE_all_HU_mid', float('nan')),
            "MAE_bone_HU_mid": test_metrics.get('MAE_bone_HU_mid', float('nan')),
            "MS_SSIM_mid": test_metrics.get('MS_SSIM_mid', float('nan')),
            "MS_SSIM_bone_mid": test_metrics.get('MS_SSIM_bone_mid', float('nan')),
            "Dice_bone_mid": test_metrics.get('Dice_bone_mid', float('nan')),
            "Jacobian_min": test_metrics.get('Jacobian_min', float('nan')),
            "Jacobian_mean": test_metrics.get('Jacobian_mean', float('nan')),
            "Jacobian_nonpos_frac": test_metrics.get('Jacobian_nonpos_frac', float('nan')),
            "Diffeomorphic": test_metrics.get('Diffeomorphic', None),
        }
        metrics_rows.append(row)

        print(f"\n🧮 Epoch {epoch:03d}/{NUM_EPOCHS} | loss={row['avg_total_loss']:.6f} | "
              f"MAE={row['MAE_all_HU']:.2f} / Bone={row['MAE_bone_HU']:.2f} / "
              f"MS-SSIM={row['MS_SSIM']:.4f} / Dice={row['Dice_bone']:.4f}")

        if HAS_TQDM:
            epoch_pbar.set_postfix({
                'loss': f"{row['avg_total_loss']:.4f}",
                'MAE_bone': f"{row['MAE_bone_HU']:.2f}",
                'Dice': f"{row['Dice_bone']:.3f}"
            })
            epoch_pbar.update(1)

        # Save checkpoint (matching FM V15 format)
        ckpt_path = CKPT_DIR / f"teacher_epoch_{epoch:04d}.pth"
        torch.save({
            'epoch': epoch,
            'model_state_dict': model.state_dict(),
            'optimizer_state_dict': optimizer.state_dict(),
            'train_loss': float(np.mean(losses)),
            'deformation_mode': DEFORMATION_MODE,
            'test_mae_all_hu': row["MAE_all_HU"],
            'test_mae_bone_hu': row["MAE_bone_HU"],
            'test_ms_ssim': row["MS_SSIM"],
            'test_dice_bone': row["Dice_bone"],
        }, ckpt_path)
        print(f"💾 Saved checkpoint: {ckpt_path}")

        # Create separate folders for train and test visualizations
        test_vis_dir = RECON_DIR / "test"
        train_vis_dir = RECON_DIR / "train"
        test_vis_dir.mkdir(parents=True, exist_ok=True)
        if SAVE_TRAIN_VIS:
            train_vis_dir.mkdir(parents=True, exist_ok=True)

        # Save visualizations for selected test samples (deterministic)
        try:
            # Gather test samples
            test_samples = []
            for b in test_loader:
                for i in range(b["x0"].size(0)):
                    test_samples.append({
                        'x0': b["x0"][i:i+1],
                        'x1': b["x1"][i:i+1],
                        'case_id': int(b["meta"]["case_id"][i]),
                        'roi_num': int(b["meta"]["roi_num"][i])
                    })
            if len(test_samples) > 0:
                vis_rng = np.random.RandomState(RANDOM_SEED)
                num_to_vis = min(NUM_VIS_SAMPLES, len(test_samples))
                idxs = vis_rng.choice(len(test_samples), size=num_to_vis, replace=False)
                sel = [test_samples[i] for i in sorted(idxs)]
                if epoch == 1:
                    case_list = [f"Patient {s['case_id']:03d} ROI {s['roi_num']:02d}" for s in sel]
                    print(f"   🎯 Test visualization samples (seed={RANDOM_SEED}): {', '.join(case_list)}")
                for vis_idx, sample in enumerate(sel):
                    x0_v = sample['x0'].to(device)
                    x1_v = sample['x1'].to(device)

                    x_hat1_v, _, _ = predict(model, x0_v, x1_v)
                    pred_hu = denorm_to_hu(x_hat1_v.squeeze(1).detach().cpu().numpy(), HU_RANGE)[0]
                    cfg_uncond_hu = None
                    cfg_cond_hu = None
                    
                    gt_hu = denorm_to_hu(x1_v.squeeze(1).detach().cpu().numpy(), HU_RANGE)[0]
                    in_hu = denorm_to_hu(x0_v.squeeze(1).detach().cpu().numpy(), HU_RANGE)[0]
                    case_id = sample['case_id']; roi_num = sample['roi_num']
                    case_info = f"Patient {case_id:03d} ROI {roi_num:02d}"

                    # Save comparison image to test folder
                    combined_path = test_vis_dir / f"epoch_{epoch:03d}_sample{vis_idx+1:02d}_case{case_id:03d}_roi{roi_num:02d}.png"
                    pred_label = "Prediction"
                    save_combined_comparison(in_hu, gt_hu, pred_hu, str(combined_path), epoch, case_info, 
                                           pred_label=pred_label, cfg_uncond_hu=cfg_uncond_hu, cfg_cond_hu=cfg_cond_hu)
                if epoch == 1:
                    print(f"   📸 Saved {num_to_vis} test visualization(s) for epoch {epoch}")
        except Exception as e:
            print(f"⚠️ Test viz save failed: {e}")

        # Save visualizations for selected train samples (deterministic, optional)
        if SAVE_TRAIN_VIS:
            try:
                # Gather train samples
                train_samples = []
                for b in train_loader:
                    for i in range(b["x0"].size(0)):
                        train_samples.append({
                            'x0': b["x0"][i:i+1],
                            'x1': b["x1"][i:i+1],
                            'case_id': int(b["meta"]["case_id"][i]),
                            'roi_num': int(b["meta"]["roi_num"][i])
                        })
                if len(train_samples) > 0:
                    vis_rng = np.random.RandomState(RANDOM_SEED)
                    num_to_vis = min(NUM_VIS_SAMPLES, len(train_samples))
                    idxs = vis_rng.choice(len(train_samples), size=num_to_vis, replace=False)
                    sel = [train_samples[i] for i in sorted(idxs)]
                    if epoch == 1:
                        case_list = [f"Patient {s['case_id']:03d} ROI {s['roi_num']:02d}" for s in sel]
                        print(f"   🎯 Train visualization samples (seed={RANDOM_SEED}): {', '.join(case_list)}")
                    for vis_idx, sample in enumerate(sel):
                        x0_v = sample['x0'].to(device)
                        x1_v = sample['x1'].to(device)

                        x_hat1_v, _, _ = predict(model, x0_v, x1_v)
                        pred_hu = denorm_to_hu(x_hat1_v.squeeze(1).detach().cpu().numpy(), HU_RANGE)[0]
                        cfg_uncond_hu = None
                        cfg_cond_hu = None
                        
                        gt_hu = denorm_to_hu(x1_v.squeeze(1).detach().cpu().numpy(), HU_RANGE)[0]
                        in_hu = denorm_to_hu(x0_v.squeeze(1).detach().cpu().numpy(), HU_RANGE)[0]
                        case_id = sample['case_id']; roi_num = sample['roi_num']
                        case_info = f"Patient {case_id:03d} ROI {roi_num:02d}"

                        # Save comparison image to train folder
                        combined_path = train_vis_dir / f"epoch_{epoch:03d}_sample{vis_idx+1:02d}_case{case_id:03d}_roi{roi_num:02d}.png"
                        pred_label = "Prediction"
                        save_combined_comparison(in_hu, gt_hu, pred_hu, str(combined_path), epoch, case_info, 
                                               pred_label=pred_label, cfg_uncond_hu=cfg_uncond_hu, cfg_cond_hu=cfg_cond_hu)
                    if epoch == 1:
                        print(f"   📸 Saved {num_to_vis} train visualization(s) for epoch {epoch}")
            except Exception as e:
                print(f"⚠️ Train viz save failed: {e}")

        # Write metrics Excel each epoch
        try:
            create_metrics_excel_with_footnotes(metrics_rows, METRICS_DIR / "training_metrics.xlsx")
        except Exception as e:
            print(f"⚠️ Failed to write metrics Excel: {e}")

    if HAS_TQDM: epoch_pbar.close()

    # Final metrics Excel (FM V18-style writer already adds AVERAGE/STD_DEV rows)
    try:
        create_metrics_excel_with_footnotes(metrics_rows, METRICS_DIR / "training_metrics.xlsx")
        print(f"📊 Metrics saved to {METRICS_DIR / 'training_metrics.xlsx'}")
    except Exception as e:
        print(f"⚠️ Failed to finalize metrics Excel: {e}")

@torch.no_grad()
def inference():
    ckpts = sorted(CKPT_DIR.glob("teacher_epoch_*.pth"))
    if not ckpts:
        print("❌ No checkpoints found for inference.")
        return
    latest = ckpts[-1]
    print(f"🔎 Loading checkpoint: {latest}")
    base_ch = UNET_BASE_CHANNELS
    model = UNet3D(base_channels=base_ch).to(device)
    data = torch.load(latest, map_location=device)
    state = data.get('state_dict')
    model.load_state_dict(state); model.eval()

    dataset = ROI3DDataset(POD5_DIR, POY1_DIR, normalize=True)
    loader = DataLoader(dataset, batch_size=BATCH_SIZE, shuffle=False, num_workers=0)

    n_batches = len(loader)
    pbar = tqdm(total=n_batches, desc="Inference", position=0) if HAS_TQDM else None

    for batch in loader:
        x0 = batch["x0"].to(device)
        x1 = batch["x1"].to(device)
        meta = batch["meta"]
        x_hat1, v_vox, phi_norm = predict(model, x0, x1)

        pred_hu = denorm_to_hu(x_hat1.squeeze(1).detach().cpu().numpy(), HU_RANGE)
        gt_hu   = denorm_to_hu(x1.squeeze(1).detach().cpu().numpy(), HU_RANGE)
        in_hu   = denorm_to_hu(x0.squeeze(1).detach().cpu().numpy(), HU_RANGE)

        for b in range(x0.size(0)):
            case_id = meta["case_id"][b]
            roi_num = meta["roi_num"][b]
            save_in = RECON_DIR / f"case_{int(case_id):03d}_roi{int(roi_num):02d}_input.png"
            save_gt = RECON_DIR / f"case_{int(case_id):03d}_roi{int(roi_num):02d}_gt.png"
            save_pr = RECON_DIR / f"case_{int(case_id):03d}_roi{int(roi_num):02d}_pred.png"
            save_orthogonal_png(in_hu[b], str(save_in), f"POD5 case {case_id} ROI {roi_num}")
            save_orthogonal_png(gt_hu[b], str(save_gt), f"POY1 GT case {case_id} ROI {roi_num}")
            save_orthogonal_png(pred_hu[b], str(save_pr), f"Prediction case {case_id} ROI {roi_num}")
        if pbar: pbar.update(1)

    if pbar: pbar.close()
    print(f"✅ Inference complete! Results saved to {RECON_DIR}")

if __name__ == "__main__":
    if MODE == 'train':
        train()
    else:
        inference()
